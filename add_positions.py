
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formula.translate import Translator
from typing import List, Tuple
import re

EXPECTED_COLS = [9,15,17,19,21,25]  # I,O,Q,S,U,Y

def find_header_and_first_data_row(ws) -> Tuple[int,int]:
    for r in range(1, ws.max_row+1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and '№ п/п' in v:
            return r, r+2
    raise RuntimeError("Не найден заголовок '№ п/п'.")

def is_sum_formula(val: str) -> bool:
    return isinstance(val, str) and val.startswith('=') and any(k in val.upper() for k in ('SUM', 'СУМ'))


def detect_first_block_subtotal(ws, first_data_row:int) -> int:
    """
    Ищем строку итогов первого блока так, чтобы она НЕ была строкой данных.
    Критерий: на строке r есть >=2 SUM в ожидаемых колонках, а на r-1 сумм НЕТ (или их меньше).
    Это защищает от ситуации, когда SUM попали в строки данных и итоги начинали "считать сами себя".
    """
    def is_sum(v):
        return isinstance(v, str) and v.startswith('=') and any(k in v.upper() for k in ('SUM','СУМ'))
    cols = [9,15,17,19,21,25]  # I,O,Q,S,U,Y

    # Основной проход
    for r in range(first_data_row+1, min(ws.max_row, first_data_row+40)+1):
        cnt = sum(1 for c in cols if is_sum(ws.cell(r,c).value))
        prev = sum(1 for c in cols if is_sum(ws.cell(r-1,c).value))
        if cnt >= 2 and prev == 0:
            return r

    # Фолбэк: первая строка ниже блока, где есть хотя бы одна SUM и выше нее нет SUM
    for r in range(first_data_row+1, min(ws.max_row, first_data_row+40)+1):
        any_sum = any(is_sum(ws.cell(r,c).value) for c in cols)
        any_sum_prev = any(is_sum(ws.cell(r-1,c).value) for c in cols)
        if any_sum and not any_sum_prev:
            return r

    # Совсем фолбэк: как раньше — первая строка с SUM в ожидаемых колонках
    for r in range(first_data_row+1, min(ws.max_row, first_data_row+40)+1):
        if any(is_sum(ws.cell(r,c).value) for c in cols):
            return r

    raise RuntimeError("Не удалось найти строку итогов первого блока.")


def copy_row_format_and_formulas(ws, src_row:int, dst_row:int):
    max_col = ws.max_column
    for c in range(1, max_col+1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        try:
            dst._style = src._style
        except Exception:
            pass
        val = src.value
        if isinstance(val, str) and val.startswith('='):
            try:
                dst.value = Translator(val, origin=f"{get_column_letter(c)}{src_row}").translate_formula(
                    f"{get_column_letter(c)}{dst_row}"
                )
            except Exception:
                dst.value = val
        else:
            dst.value = val


def write_sums(ws, subtotal_row:int, first_data_row:int):
    last_data_row = subtotal_row - 1
    for c in EXPECTED_COLS:
        # 1) Проверим наличие данных в столбце
        has_data = False
        for r in range(first_data_row, last_data_row+1):
            v = ws.cell(r,c).value
            if v not in (None, ''):
                has_data = True
                break
        if not has_data:
            continue

        # 2) Если ячейка subtotal находится внутри merged-диапазона — писать формулу нужно в верхний левый угол
        target_col = c
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= subtotal_row <= rng.max_row and rng.min_col <= c <= rng.max_col:
                target_col = rng.min_col
                break

        col_letter = get_column_letter(c)
        ws.cell(subtotal_row, target_col).value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"

def write_avg_x(ws, subtotal_row:int, first_data_row:int):
    """В столбце X (24) в строке итогов поставить среднее по блоку позиций: =AVERAGE(Xfirst:Xlast).
    Учитываем слияния: пишем в верхний левый угол, если X попал в merged-диапазон строки итогов.
    """
    from openpyxl.utils import get_column_letter
    X = 24  # колонка X
    last_data_row = subtotal_row - 1

    # Проверим наличие данных в X в блоке
    has_data = False
    for r in range(first_data_row, last_data_row+1):
        if ws.cell(r, X).value not in (None, ''):
            has_data = True
            break
    if not has_data:
        return

    # Куда писать: если X в merge на строке итогов — используем верхний левый столбец merge
    target_col = X
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= subtotal_row <= rng.max_row and rng.min_col <= X <= rng.max_col:
            target_col = rng.min_col
            break

    col_letter = get_column_letter(X)
    ws.cell(subtotal_row, target_col).value = f"=AVERAGE({col_letter}{first_data_row}:{col_letter}{last_data_row})"


def shift_row_refs_in_formula(v:str, threshold:int, delta:int) -> str:
    """Сдвигаем ссылки на строки >= threshold на delta (и для абсолютных, и для относительных)."""
    if not (isinstance(v,str) and v.startswith('=')):
        return v
    def repl(m):
        col = m.group(1)
        row = int(m.group(2))
        if row >= threshold:
            return f"{col}{row+delta}"
        return m.group(0)
    return re.sub(r'(\$?[A-Z]+\$?)(\d+)', repl, v)

def shift_row_refs_in_block(ws, first_data_row:int, subtotal_row_old:int, how_many:int):
    """Проходим по строкам позиций и сдвигаем ссылки, чтобы учесть вставку строк над итогами."""
    for r in range(first_data_row, subtotal_row_old):
        for c in range(1, ws.max_column+1):
            v = ws.cell(r,c).value
            if isinstance(v,str) and v.startswith('='):
                ws.cell(r,c).value = shift_row_refs_in_formula(v, subtotal_row_old, how_many)

def retarget_row_refs(v:str, old_row:int, new_row:int) -> str:
    if not (isinstance(v,str) and v.startswith('=')):
        return v
    def repl(m):
        col = m.group(1)
        row = int(m.group(2))
        if row == old_row:
            return f"{col}{new_row}"
        return m.group(0)
    return re.sub(r'(\$?[A-Z]+\$?)(\d+)', repl, v)

def update_row_refs(ws, row:int, old_row:int, new_row:int):
    for c in range(1, ws.max_column+1):
        v = ws.cell(row,c).value
        if isinstance(v,str) and v.startswith('='):
            ws.cell(row,c).value = retarget_row_refs(v, old_row, new_row)
def update_logistics_formulas(ws, first_data_row:int, subtotal_row:int, u_anchor_row:int=15):
    """
    Перезаписывает формулы логистики для столбцов R (18) и T (20) в строках позиций:
      Rr := =$U$<u_anchor_row>/Q<subtotal_row>*P<r>/12*0.3
      Tr := =$U$<u_anchor_row>/Q<subtotal_row>*P<r>/12*0.7
    где Q<subtotal_row> — общий вес блока (итог по колонке Q).
    """
    for r in range(first_data_row, subtotal_row):
        ws.cell(r, 18).value = f"=$U${u_anchor_row}/Q{subtotal_row}*P{r}/12*0.3"
        ws.cell(r, 20).value = f"=$U${u_anchor_row}/Q{subtotal_row}*P{r}/12*0.7"


def add_positions(src_path:str, dst_path:str, sheet_name:str='расчет', how_many:int=1, u_anchor_row:int=15) -> None:
    wb = load_workbook(src_path, data_only=False)
    ws = wb[sheet_name]

    header_row, first_data_row = find_header_and_first_data_row(ws)
    old_subtotal_row = detect_first_block_subtotal(ws, first_data_row)

    template_row = first_data_row
    template_height = ws.row_dimensions[template_row].height

    # Сдвигаем ссылки в формулах строк позиций, чтобы Excel-формулы продолжали ссылаться на правильные строки после вставки
    shift_row_refs_in_block(ws, first_data_row, old_subtotal_row, how_many)

    insert_at = old_subtotal_row
    for _ in range(how_many):
        ws.insert_rows(insert_at)
        copy_row_format_and_formulas(ws, template_row, insert_at)
        ws.cell(insert_at, 2).value = None
        if template_height is not None:
            ws.row_dimensions[insert_at].height = template_height
        insert_at += 1

    new_subtotal_row = old_subtotal_row + how_many
    write_sums(ws, new_subtotal_row, first_data_row)
    write_avg_x(ws, new_subtotal_row, first_data_row)
    update_row_refs(ws, new_subtotal_row, old_subtotal_row, new_subtotal_row)
    update_row_refs(ws, new_subtotal_row+1, old_subtotal_row, new_subtotal_row)

    update_logistics_formulas(ws, first_data_row, new_subtotal_row, u_anchor_row)

    wb.save(dst_path)

if __name__ == '__main__':
    import argparse
    ap = argparse.ArgumentParser(description='Добавить позиции в бюджетный шаблон, сохраняя формулы и форматы (версия под ГП521452).')
    ap.add_argument('--src', required=True, help='Путь к исходному .xlsx')
    ap.add_argument('--dst', required=True, help='Путь для сохранения результата .xlsx')
    ap.add_argument('--sheet', default='расчет', help='Имя листа (по умолчанию "расчет")')
    ap.add_argument('--n', type=int, default=1, help='Сколько новых позиций добавить (по умолчанию 1)')
    ap.add_argument('--u_anchor_row', type=int, default=15, help='Номер строки для $U$<row> в формулах логистики (по умолчанию 15)')
    args = ap.parse_args()
    add_positions(args.src, args.dst, args.sheet, args.n, args.u_anchor_row)
