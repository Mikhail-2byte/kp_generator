from __future__ import annotations

import csv
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional
from uuid import uuid4

from flask import current_app

from app.core.cache import cached_dataset


BASE_DIR = Path(__file__).resolve().parents[2]
CONFIG_DIR = BASE_DIR / 'config'
VERSIONS_DIR = CONFIG_DIR / 'versions'

GB_ANALOGS: List[Dict[str, Any]] = []
DUTY_RATES: List[Dict[str, Any]] = []
TNVED_CATALOG: List[Dict[str, Any]] = []
ORDERS_REGISTRY: List[Dict[str, Any]] = []
TASK_TEMPLATES: List[Dict[str, Any]] = []
TASK_INSTRUCTIONS: List[Dict[str, Any]] = []


def _log_error(message: str):
    """Пишет сообщение об ошибке в лог приложения, если он доступен."""
    logger = getattr(current_app, 'logger', None)
    if logger:
        logger.error(message)


@cached_dataset(maxsize=1)
def load_gb_materials() -> List[Dict[str, Any]]:
    """Читает аналоги материалов из конфигурационного JSON."""
    materials_path = CONFIG_DIR / 'gb_materials.json'
    try:
        with materials_path.open('r', encoding='utf-8') as file:
            data = json.load(file)
        materials = data.get('materials', [])
        return materials
    except FileNotFoundError:
        _log_error(f'GB materials file not found at {materials_path.as_posix()}')
    except json.JSONDecodeError as exc:
        _log_error(f'Failed to parse GB materials file: {exc}')
    return []


def save_gb_materials(materials: List[Dict[str, Any]], *, actor: Optional[str] = None) -> None:
    """Сохраняет список аналогов материалов обратно в файл."""
    materials_path = CONFIG_DIR / 'gb_materials.json'
    materials_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        'materials': [
            {
                'russian': material.get('russian', ''),
                'gb': material.get('gb', ''),
                'notes': material.get('notes', ''),
                'gost': material.get('gost', ''),
                'price': material.get('price', ''),
                'workpiece_type': material.get('workpiece_type', '')
            }
            for material in materials
        ]
    }
    _snapshot_version('gb_materials', payload, actor)

    with materials_path.open('w', encoding='utf-8') as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)
    
    # Инвалидируем кэш после сохранения
    load_gb_materials.cache_clear()  # type: ignore


def refresh_gb_analogs() -> None:
    """Перечитывает аналоги материалов в память для быстрого доступа."""
    global GB_ANALOGS
    GB_ANALOGS = load_gb_materials()


def get_gb_materials() -> List[Dict[str, Any]]:
    """Возвращает копию кэшированного списка аналогов материалов."""
    return list(GB_ANALOGS)


@cached_dataset(maxsize=1)
def load_duty_rates() -> List[Dict[str, Any]]:
    """Загружает ставки пошлин из tnved_catalog.json и готовит поля для поиска.
    
    Все пошлины хранятся в едином файле tnved_catalog.json.
    Поддерживает два типа записей:
    - Простые: product, category, duty_percent
    - Расширенные (ТН ВЭД): code, description, keywords_display, examples, duty_text, duty_percent
    """
    duty_path = CONFIG_DIR / 'tnved_catalog.json'
    try:
        with duty_path.open('r', encoding='utf-8') as file:
            data = json.load(file)
        items = data.get('items', [])

        for item in items:
            # Определяем тип записи
            if 'code' in item and item.get('code'):
                # Расширенная запись (ТН ВЭД)
                code = item.get('code', '')
                description = item.get('description', '')
                keywords_text = item.get('keywords_display', '')
                examples_text = item.get('examples', '')
                duty_text = item.get('duty_text', '') or '—'
                duty_percent = item.get('duty_percent')
                
                # Если duty_percent не указан, пытаемся извлечь из duty_text
                if duty_percent is None:
                    duty_percent = _extract_percent_value(duty_text)
                
                keywords_list = _split_keywords(keywords_text)
                
                search_chunks = [
                    code.lower(),
                    keywords_text.lower(),
                    description.lower(),
                    examples_text.lower(),
                    duty_text.lower(),
                ]
                
                # Добавляем поля для совместимости
                item.update({
                    'keywords_list': keywords_list,
                    'title': keywords_text or description or code,
                    'title_search': (keywords_text or description or code).lower(),
                    'description_search': description.lower(),
                    'examples_search': examples_text.lower(),
                    'duty_search': duty_text.lower(),
                    'search_blob': ' '.join(filter(None, search_chunks)),
                    'source': 'tnved',
                    # Для обратной совместимости с get_duty_catalog
                    'product': keywords_text or description or code,
                    'category': description,
                    'product_search': (keywords_text or description or code).lower(),
                    'category_search': description.lower(),
                })
            else:
                # Простая запись
                item['product_search'] = str(item.get('product', '')).lower()
                item['category_search'] = str(item.get('category', '')).lower()
                item['duty_search'] = str(item.get('duty_percent', '')).lower()
                item['source'] = 'manual'

        return items
    except FileNotFoundError:
        _log_error(f'Duty rates file not found at {duty_path.as_posix()}')
    except json.JSONDecodeError as exc:
        _log_error(f'Failed to parse duty rates file: {exc}')
    return []


def _tnved_catalog_path() -> Optional[Path]:
    """Определяет путь к CSV каталогу ТН ВЭД.

    Предпочтительно используем файл из статических ресурсов проекта,
    если он присутствует, чтобы не требовать ручного копирования в корень.
    """
    data_dir = BASE_DIR / 'data'
    static_dir = BASE_DIR / 'static'
    candidates = [
        # Основные, “официальные” пути
        data_dir / 'tnved_catalog.csv',
        CONFIG_DIR / 'tnved_catalog.csv',
        # Файл, который обычно лежит в репозитории (static/)
        static_dir / 'ТН-ВЭД-ТД-для-менеджеров-с-ключевыми-словами.csv',
        # Исторический путь в корне проекта
        BASE_DIR / 'ТН-ВЭД-ТД-для-менеджеров-с-ключевыми-словами.csv',
    ]

    for candidate in candidates:
        if candidate.exists():
            return candidate

    # На всякий случай пытаемся найти файл по маске, включая подкаталоги static/
    try:
        for path in BASE_DIR.rglob('*ТН*ВЭД*csv'):
            if path.is_file():
                return path
    except OSError:
        return None

    return None


def _split_keywords(raw_text: str) -> List[str]:
    """Разбивает строку ключевых слов на список значений."""
    if not raw_text:
        return []
    parts = re.split(r'[;,/]|(?:\s{2,})', raw_text)
    keywords = []
    for part in parts:
        cleaned = part.strip()
        if cleaned:
            keywords.append(cleaned)
    if not keywords and raw_text.strip():
        keywords.append(raw_text.strip())
    return keywords


def _extract_percent_value(raw_text: str) -> Optional[float]:
    """Извлекает числовое значение процента из текстового описания."""
    if not raw_text:
        return None
    match = re.search(r'(\d+(?:[.,]\d+)?)\s*%', raw_text)
    if not match:
        return None
    value = match.group(1).replace(',', '.')
    try:
        return float(value)
    except ValueError:
        return None


def _format_percent_display(value: Optional[float]) -> str:
    """Формирует строковое представление значения процента."""
    if value is None:
        return '—'
    formatted = f'{value:.2f}'.rstrip('0').rstrip('.')
    return f'{formatted}%'


@cached_dataset(maxsize=1)
def load_tnved_catalog() -> List[Dict[str, Any]]:
    """Загружает расширенный каталог ставок пошлин из JSON (основной источник) или CSV (fallback).
    
    Читает из единого файла tnved_catalog.json и возвращает только записи с полем 'code' (ТН ВЭД).
    """
    tnved_json_path = CONFIG_DIR / 'tnved_catalog.json'
    
    # Пытаемся загрузить из JSON (основной источник)
    if tnved_json_path.exists():
        try:
            with tnved_json_path.open('r', encoding='utf-8') as file:
                data = json.load(file)
            items_data = data.get('items', [])
            
            items: List[Dict[str, Any]] = []
            for item_data in items_data:
                # Пропускаем простые записи (без code)
                if 'code' not in item_data or not item_data.get('code'):
                    continue
                    
                code = item_data.get('code', '').strip()
                description = item_data.get('description', '').strip()
                keywords_text = item_data.get('keywords_display', '').strip()
                examples_text = item_data.get('examples', '').strip()
                duty_text = item_data.get('duty_text', '').strip() or '—'
                duty_percent = item_data.get('duty_percent')
                
                # Если duty_percent не указан, пытаемся извлечь из duty_text
                if duty_percent is None:
                    duty_percent = _extract_percent_value(duty_text)
                
                keywords_list = _split_keywords(keywords_text)
                
                search_chunks = [
                    code.lower(),
                    keywords_text.lower(),
                    description.lower(),
                    examples_text.lower(),
                    duty_text.lower(),
                ]
                
                items.append({
                    'code': code,
                    'description': description,
                    'keywords_display': keywords_text,
                    'keywords_list': keywords_list,
                    'examples': examples_text,
                    'duty_text': duty_text,
                    'duty_percent': duty_percent,
                    'title': keywords_text or description or code,
                    'title_search': (keywords_text or description or code).lower(),
                    'description_search': description.lower(),
                    'examples_search': examples_text.lower(),
                    'duty_search': duty_text.lower(),
                    'search_blob': ' '.join(filter(None, search_chunks)),
                    'source': 'tnved',
                })
            
            return items
        except (FileNotFoundError, json.JSONDecodeError) as exc:
            _log_error(f'Failed to load TNVED catalog from JSON: {exc}')
    
    # Fallback: загружаем из CSV (для обратной совместимости)
    catalog_path = _tnved_catalog_path()
    if catalog_path is None:
        _log_error('TNVED catalog file not found.')
        return []

    items: List[Dict[str, Any]] = []
    try:
        with catalog_path.open('r', encoding='utf-8-sig', newline='') as csv_file:
            reader = csv.reader(csv_file)
            for row in reader:
                if not row or not any(cell.strip() for cell in row):
                    continue

                if len(row) < 6:
                    continue

                index_raw = row[0].strip().strip('"')
                if not index_raw or not index_raw.isdigit():
                    continue

                code = row[1].strip()
                description = row[2].strip()
                keywords_text = row[3].strip()
                examples_text = row[4].strip()
                duty_text = row[5].strip() or '—'

                keywords_list = _split_keywords(keywords_text)
                duty_percent = _extract_percent_value(duty_text)

                search_chunks = [
                    code.lower(),
                    keywords_text.lower(),
                    description.lower(),
                    examples_text.lower(),
                    duty_text.lower(),
                ]

                items.append({
                    'code': code,
                    'description': description,
                    'keywords_display': keywords_text,
                    'keywords_list': keywords_list,
                    'examples': examples_text,
                    'duty_text': duty_text,
                    'duty_percent': duty_percent,
                    'title': keywords_text or description or code,
                    'title_search': (keywords_text or description or code).lower(),
                    'description_search': description.lower(),
                    'examples_search': examples_text.lower(),
                    'duty_search': duty_text.lower(),
                    'search_blob': ' '.join(filter(None, search_chunks)),
                    'source': 'tnved',
                })
    except FileNotFoundError:
        _log_error(f'TNVED catalog file not found: {catalog_path}')
    except Exception as exc:  # pragma: no cover - defensive logging
        _log_error(f'Failed to parse TNVED catalog: {exc}')

    return items


def save_tnved_catalog(items: List[Dict[str, Any]], *, actor: Optional[str] = None) -> None:
    """Сохраняет каталог ТН ВЭД в JSON файл."""
    tnved_path = CONFIG_DIR / 'tnved_catalog.json'
    tnved_path.parent.mkdir(parents=True, exist_ok=True)
    
    def _coerce_float(value):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    
    payload = {
        'items': [
            {
                'code': item.get('code', ''),
                'description': item.get('description', ''),
                'keywords_display': item.get('keywords_display', ''),
                'examples': item.get('examples', ''),
                'duty_text': item.get('duty_text', ''),
                'duty_percent': _coerce_float(item.get('duty_percent'))
            }
            for item in items
        ]
    }
    
    _snapshot_version('tnved_catalog', payload, actor)
    
    with tnved_path.open('w', encoding='utf-8') as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)
    
    # Инвалидируем кэш после сохранения
    load_tnved_catalog.cache_clear()  # type: ignore


def save_duty_rates(items: List[Dict[str, Any]], *, actor: Optional[str] = None) -> None:
    """Сохраняет изменённый список ставок пошлин в tnved_catalog.json.
    
    Все пошлины хранятся в едином файле tnved_catalog.json.
    Поддерживает два типа записей:
    - Простые: product, category, duty_percent
    - Расширенные (ТН ВЭД): code, description, keywords_display, examples, duty_text, duty_percent
    """
    duty_path = CONFIG_DIR / 'tnved_catalog.json'
    duty_path.parent.mkdir(parents=True, exist_ok=True)

    def _coerce_float(value):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    payload_items = []
    for item in items:
        # Определяем тип записи: если есть code - это ТН ВЭД, иначе простая запись
        if 'code' in item and item.get('code'):
            # Расширенная запись (ТН ВЭД)
            payload_items.append({
                'code': item.get('code', ''),
                'description': item.get('description', ''),
                'keywords_display': item.get('keywords_display', ''),
                'examples': item.get('examples', ''),
                'duty_text': item.get('duty_text', ''),
                'duty_percent': _coerce_float(item.get('duty_percent'))
            })
        else:
            # Простая запись
            payload_items.append({
                'product': item.get('product', ''),
                'category': item.get('category', ''),
                'duty_percent': _coerce_float(item.get('duty_percent', 0)) or 0.0
            })

    payload = {'items': payload_items}

    _snapshot_version('tnved_catalog', payload, actor)

    with duty_path.open('w', encoding='utf-8') as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)
    
    # Инвалидируем кэш после сохранения
    load_duty_rates.cache_clear()  # type: ignore
    load_tnved_catalog.cache_clear()  # type: ignore


def refresh_duty_rates() -> None:
    """Обновляет кэш ставок пошлин после изменения файлов."""
    global DUTY_RATES
    # Очищаем кэш перед загрузкой
    load_duty_rates.cache_clear()  # type: ignore
    DUTY_RATES = load_duty_rates()


def get_duty_rates() -> List[Dict[str, Any]]:
    """Возвращает копию кэшированного списка ставок пошлин."""
    return list(DUTY_RATES)


def refresh_tnved_catalog() -> None:
    """Обновляет кэш расширенного каталога ставок пошлин."""
    global TNVED_CATALOG
    # Очищаем кэш перед загрузкой
    load_tnved_catalog.cache_clear()  # type: ignore
    TNVED_CATALOG = load_tnved_catalog()


def get_tnved_catalog() -> List[Dict[str, Any]]:
    """Возвращает копию каталога ставок из CSV."""
    return list(TNVED_CATALOG)


def _normalize_manual_duty_item(item: Dict[str, Any]) -> Dict[str, Any]:
    """Нормализует одну простую запись пошлины для использования в get_duty_catalog."""
    return {
        **item,
        'product': item.get('product', ''),
        'category': item.get('category', ''),
        'source': 'manual'
    }


def _normalize_manual_duty_items(manual_items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Преобразует ручные записи пошлин в общий формат каталога."""
    normalized: List[Dict[str, Any]] = []
    for item in manual_items:
        product = str(item.get('product', '')).strip()
        category = str(item.get('category', '')).strip()
        duty_percent = item.get('duty_percent')
        duty_text = _format_percent_display(duty_percent if duty_percent is not None else None)
        title = product or category or 'Без названия'
        title_search = title.lower()
        description_search = category.lower()
        duty_search = duty_text.lower()

        normalized.append({
            'code': '',
            'description': category,
            'keywords_display': product,
            'keywords_list': [product] if product else [],
            'examples': '',
            'duty_text': duty_text,
            'duty_percent': duty_percent,
            'title': title,
            'title_search': title_search,
            'description_search': description_search,
            'examples_search': '',
            'duty_search': duty_search,
            'search_blob': ' '.join(filter(None, [title_search, description_search, duty_search])),
            'product': product,
            'category': category,
            'source': 'manual',
        })
    return normalized


def _normalize_tnved_items(catalog_items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Нормализует записи каталога ТН ВЭД (здесь данные уже стандартизированы)."""
    normalized: List[Dict[str, Any]] = []
    for item in catalog_items:
        normalized.append({
            **item,
            'product': item.get('keywords_display', ''),
            'category': item.get('description', ''),
        })
    return normalized


def _duty_catalog_sort_key(item: Dict[str, Any]) -> tuple:
    """Используется для устойчивой сортировки общего каталога."""
    source_priority = 0 if item.get('source') == 'manual' else 1
    code = item.get('code') or ''
    title = item.get('title') or ''
    return (source_priority, code, title.lower())


def get_duty_catalog() -> List[Dict[str, Any]]:
    """Возвращает каталог пошлин из единого файла duty_rates.json.
    
    Все записи (и простые, и ТН ВЭД) теперь хранятся в одном файле.
    """
    items = load_duty_rates()
    # Нормализуем записи для совместимости с существующим API
    normalized = []
    for item in items:
        if item.get('code'):
            # Расширенная запись (ТН ВЭД) - уже нормализована в load_duty_rates
            normalized.append(item)
        else:
            # Простая запись - нормализуем через существующую функцию
            normalized_items = _normalize_manual_duty_items([item])
            if normalized_items:
                normalized.append(normalized_items[0])
    return sorted(normalized, key=_duty_catalog_sort_key)


@cached_dataset(maxsize=1)
def load_logistics_cities() -> List[Dict[str, Any]]:
    """
    Загружает справочник городов и тарифов логистики.
    Объединяет все три справочника (основные города, ЕКБ+РФ, трал).
    """
    # Всегда используем объединение новых справочников
    # Старый файл logistics_cities.json больше не используется напрямую
    return load_all_logistics_cities()


def save_logistics_cities(cities: List[Dict[str, Any]], *, actor: Optional[str] = None) -> None:
    """Сохраняет обновлённый список тарифов логистики в JSON."""
    logistics_path = CONFIG_DIR / 'logistics_cities.json'
    logistics_path.parent.mkdir(parents=True, exist_ok=True)

    def _coerce_float(value):
        try:
            return float(value) if value is not None else None
        except (TypeError, ValueError):
            return None
    
    def _coerce_int(value):
        try:
            return int(value) if value is not None else None
        except (TypeError, ValueError):
            return None

    payload = {
        'cities': [
            {
                'name': city.get('name', ''),
                'region': city.get('region', ''),
                'truck_price': _coerce_float(city.get('truck_price', 0)),
                'trail_price': _coerce_float(city.get('trail_price')),
                'is_main_route': bool(city.get('is_main_route', False)),
                'allows_trail': bool(city.get('allows_trail', False)),
                'distance_from_ekb_km': _coerce_int(city.get('distance_from_ekb_km')),
                **({'main_city': city['main_city']} if 'main_city' in city else {})
            }
            for city in cities
        ]
    }

    _snapshot_version('logistics_cities', payload, actor)

    with logistics_path.open('w', encoding='utf-8') as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)
    
    # Инвалидируем кэш после сохранения
    load_logistics_cities.cache_clear()  # type: ignore
    load_all_logistics_cities.cache_clear()  # type: ignore


@cached_dataset(maxsize=1)
def load_main_cities() -> List[Dict[str, Any]]:
    """
    Загружает справочник основных городов и городов в радиусе 300км.
    Включает основные города (is_main_route=True) и города с полем main_city.
    """
    main_cities_path = CONFIG_DIR / 'logistics_main_cities.json'
    try:
        with main_cities_path.open('r', encoding='utf-8') as file:
            data = json.load(file)
        return data.get('cities', [])
    except FileNotFoundError:
        _log_error(f'Main cities file not found at {main_cities_path.as_posix()}')
    except json.JSONDecodeError as exc:
        _log_error(f'Failed to parse main cities file: {exc}')
    return []


def save_main_cities(cities: List[Dict[str, Any]], *, actor: Optional[str] = None) -> None:
    """Сохраняет справочник основных городов и городов в радиусе 300км."""
    main_cities_path = CONFIG_DIR / 'logistics_main_cities.json'
    main_cities_path.parent.mkdir(parents=True, exist_ok=True)

    def _coerce_float(value):
        try:
            return float(value) if value is not None else None
        except (TypeError, ValueError):
            return None

    payload = {
        'cities': [
            {
                'name': city.get('name', ''),
                'region': city.get('region', ''),
                'truck_price': _coerce_float(city.get('truck_price', 0)),
                'is_main_route': bool(city.get('is_main_route', False)),
                'main_city': city.get('main_city')  # Для городов в радиусе 300км
            }
            for city in cities
        ]
    }

    _snapshot_version('logistics_main_cities', payload, actor)

    with main_cities_path.open('w', encoding='utf-8') as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)
    
    # Инвалидируем кэш после сохранения
    load_main_cities.cache_clear()  # type: ignore
    load_all_logistics_cities.cache_clear()  # type: ignore
    load_logistics_cities.cache_clear()  # type: ignore


@cached_dataset(maxsize=1)
def load_ekb_rf_cities() -> List[Dict[str, Any]]:
    """
    Загружает справочник городов за пределами 300км (используется алгоритм ЕКБ+РФ).
    """
    ekb_rf_path = CONFIG_DIR / 'logistics_ekb_rf_cities.json'
    try:
        with ekb_rf_path.open('r', encoding='utf-8') as file:
            data = json.load(file)
        return data.get('cities', [])
    except FileNotFoundError:
        _log_error(f'EKB+RF cities file not found at {ekb_rf_path.as_posix()}')
    except json.JSONDecodeError as exc:
        _log_error(f'Failed to parse EKB+RF cities file: {exc}')
    return []


def save_ekb_rf_cities(cities: List[Dict[str, Any]], *, actor: Optional[str] = None) -> None:
    """Сохраняет справочник городов для алгоритма ЕКБ+РФ."""
    ekb_rf_path = CONFIG_DIR / 'logistics_ekb_rf_cities.json'
    ekb_rf_path.parent.mkdir(parents=True, exist_ok=True)

    def _coerce_int(value):
        try:
            return int(value) if value is not None else None
        except (TypeError, ValueError):
            return None

    payload = {
        'cities': [
            {
                'name': city.get('name', ''),
                'region': city.get('region', ''),
                'distance_from_ekb_km': _coerce_int(city.get('distance_from_ekb_km'))
            }
            for city in cities
        ]
    }

    _snapshot_version('logistics_ekb_rf_cities', payload, actor)

    with ekb_rf_path.open('w', encoding='utf-8') as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)
    
    # Инвалидируем кэш после сохранения
    load_ekb_rf_cities.cache_clear()  # type: ignore
    load_all_logistics_cities.cache_clear()  # type: ignore
    load_logistics_cities.cache_clear()  # type: ignore


@cached_dataset(maxsize=1)
def load_trail_cities() -> List[Dict[str, Any]]:
    """Загружает справочник городов для трала."""
    trail_path = CONFIG_DIR / 'logistics_trail_cities.json'
    try:
        with trail_path.open('r', encoding='utf-8') as file:
            data = json.load(file)
        return data.get('cities', [])
    except FileNotFoundError:
        _log_error(f'Trail cities file not found at {trail_path.as_posix()}')
    except json.JSONDecodeError as exc:
        _log_error(f'Failed to parse trail cities file: {exc}')
    return []


def save_trail_cities(cities: List[Dict[str, Any]], *, actor: Optional[str] = None) -> None:
    """Сохраняет справочник городов для трала."""
    trail_path = CONFIG_DIR / 'logistics_trail_cities.json'
    trail_path.parent.mkdir(parents=True, exist_ok=True)

    def _coerce_float(value):
        try:
            return float(value) if value is not None else None
        except (TypeError, ValueError):
            return None

    payload = {
        'cities': [
            {
                'name': city.get('name', ''),
                'region': city.get('region', ''),
                'trail_price': _coerce_float(city.get('trail_price', 0))
            }
            for city in cities
        ]
    }

    _snapshot_version('logistics_trail_cities', payload, actor)

    with trail_path.open('w', encoding='utf-8') as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)
    
    # Инвалидируем кэш после сохранения
    load_trail_cities.cache_clear()  # type: ignore
    load_all_logistics_cities.cache_clear()  # type: ignore
    load_logistics_cities.cache_clear()  # type: ignore


def export_main_cities_to_excel() -> bytes:
    """Экспортирует основные города в Excel формат.
    
    Returns:
        Байты Excel файла
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError as exc:
        raise RuntimeError('openpyxl не установлен. Установите: pip install openpyxl') from exc
    
    from io import BytesIO
    
    cities = load_main_cities()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Основные города'
    
    # Стили для заголовков
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # Заголовки
    headers = ['Город', 'Регион', 'Фура, руб.', 'Основной маршрут', 'Основной город']
    ws.append(headers)
    
    # Применяем стили к заголовкам
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Данные
    for city in cities:
        ws.append([
            city.get('name', ''),
            city.get('region', ''),
            city.get('truck_price', 0) or 0,
            'Да' if city.get('is_main_route') else 'Нет',
            city.get('main_city', '') or ''
        ])
    
    # Автоподбор ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (AttributeError, TypeError):
                pass  # Игнорируем ошибки преобразования значений ячеек
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в байты
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


def import_main_cities_from_excel(excel_path: Path, *, actor: Optional[str] = None) -> int:
    """Импортирует основные города из Excel файла, заменяя все существующие данные.
    
    Args:
        excel_path: Путь к Excel файлу
        actor: Имя пользователя, выполняющего импорт (для версионирования)
        
    Returns:
        Количество импортированных записей
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError('openpyxl не установлен. Установите: pip install openpyxl') from exc
    
    if not excel_path.exists():
        raise FileNotFoundError(f'Файл не найден: {excel_path}')
    
    cities = []
    
    wb = None
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Пропускаем заголовок (первая строка)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        for row in rows:
            if not row or len(row) < 3:
                continue
            
            # Пропускаем полностью пустые строки
            if not any(cell for cell in row if cell is not None and str(cell).strip()):
                continue
            
            # Индексы колонок: 0-Город, 1-Регион, 2-Фура, руб., 3-Основной маршрут, 4-Основной город
            name = str(row[0]).strip() if row[0] else ''
            region = str(row[1]).strip() if row[1] else ''
            truck_price_value = row[2] if len(row) > 2 else None
            is_main_route_value = row[3] if len(row) > 3 else None
            main_city = str(row[4]).strip() if len(row) > 4 and row[4] else None
            
            # Преобразуем цену фуры в число
            try:
                if truck_price_value is None:
                    truck_price = 0.0
                elif isinstance(truck_price_value, (int, float)):
                    truck_price = float(truck_price_value)
                else:
                    truck_price = float(str(truck_price_value).strip().replace(',', '.'))
            except (ValueError, TypeError):
                truck_price = 0.0
            
            # Преобразуем флаг основного маршрута
            is_main_route = False
            if is_main_route_value is not None:
                if isinstance(is_main_route_value, bool):
                    is_main_route = is_main_route_value
                elif isinstance(is_main_route_value, str):
                    is_main_route = is_main_route_value.strip().lower() in ('да', 'yes', 'true', '1', 'y')
                elif isinstance(is_main_route_value, (int, float)):
                    is_main_route = bool(is_main_route_value)
            
            if name:
                cities.append({
                    'name': name,
                    'region': region,
                    'truck_price': truck_price,
                    'is_main_route': is_main_route,
                    'main_city': main_city if main_city else None
                })
        
    except Exception as exc:
        try:
            current_app.logger.error(f'Ошибка при парсинге Excel: {exc}')
        except RuntimeError:
            pass
        raise
    finally:
        # Гарантированно закрываем файл перед удалением
        if wb is not None:
            wb.close()
    
    # Защита от удаления всех городов: сохраняем только если есть хотя бы одна валидная запись
    if cities:
        save_main_cities(cities, actor=actor)
        load_main_cities.cache_clear()  # type: ignore
        load_all_logistics_cities.cache_clear()  # type: ignore
        load_logistics_cities.cache_clear()  # type: ignore
    elif rows_count > 0:
        # Если были строки в файле, но ни одна не была валидной - это ошибка
        raise ValueError(
            'Не удалось импортировать ни одной записи. '
            'Проверьте, что файл содержит данные и все обязательные поля заполнены.'
        )
    
    return len(cities)


def export_ekb_rf_cities_to_excel() -> bytes:
    """Экспортирует города ЕКБ+РФ в Excel формат.
    
    Returns:
        Байты Excel файла
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError as exc:
        raise RuntimeError('openpyxl не установлен. Установите: pip install openpyxl') from exc
    
    from io import BytesIO
    
    cities = load_ekb_rf_cities()
    wb = Workbook()
    ws = wb.active
    ws.title = 'ЕКБ+РФ города'
    
    # Стили для заголовков
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # Заголовки
    headers = ['Город', 'Регион', 'Расстояние от ЕКБ, км']
    ws.append(headers)
    
    # Применяем стили к заголовкам
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Данные
    for city in cities:
        ws.append([
            city.get('name', ''),
            city.get('region', ''),
            city.get('distance_from_ekb_km') if city.get('distance_from_ekb_km') is not None else ''
        ])
    
    # Автоподбор ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (AttributeError, TypeError):
                pass  # Игнорируем ошибки преобразования значений ячеек
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в байты
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


def import_ekb_rf_cities_from_excel(excel_path: Path, *, actor: Optional[str] = None) -> int:
    """Импортирует города ЕКБ+РФ из Excel файла, заменяя все существующие данные.
    
    Args:
        excel_path: Путь к Excel файлу
        actor: Имя пользователя, выполняющего импорт (для версионирования)
        
    Returns:
        Количество импортированных записей
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError('openpyxl не установлен. Установите: pip install openpyxl') from exc
    
    if not excel_path.exists():
        raise FileNotFoundError(f'Файл не найден: {excel_path}')
    
    cities = []
    rows_count = 0
    
    wb = None
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Пропускаем заголовок (первая строка)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        rows_count = len(rows)
        
        for row in rows:
            if not row or len(row) < 2:
                continue
            
            # Пропускаем полностью пустые строки
            if not any(cell for cell in row if cell is not None and str(cell).strip()):
                continue
            
            # Индексы колонок: 0-Город, 1-Регион, 2-Расстояние от ЕКБ, км
            name = str(row[0]).strip() if row[0] else ''
            region = str(row[1]).strip() if row[1] else ''
            distance_value = row[2] if len(row) > 2 else None
            
            # Преобразуем расстояние в число (может быть пустым)
            distance_from_ekb_km = None
            if distance_value is not None:
                try:
                    if isinstance(distance_value, (int, float)):
                        distance_from_ekb_km = int(distance_value)
                    else:
                        distance_str = str(distance_value).strip()
                        if distance_str:
                            distance_from_ekb_km = int(float(distance_str.replace(',', '.')))
                except (ValueError, TypeError):
                    distance_from_ekb_km = None
            
            if name:
                cities.append({
                    'name': name,
                    'region': region,
                    'distance_from_ekb_km': distance_from_ekb_km
                })
        
    except Exception as exc:
        try:
            current_app.logger.error(f'Ошибка при парсинге Excel: {exc}')
        except RuntimeError:
            pass
        raise
    finally:
        # Гарантированно закрываем файл перед удалением
        if wb is not None:
            wb.close()
    
    # Защита от удаления всех городов: сохраняем только если есть хотя бы одна валидная запись
    if cities:
        save_ekb_rf_cities(cities, actor=actor)
        load_ekb_rf_cities.cache_clear()  # type: ignore
        load_all_logistics_cities.cache_clear()  # type: ignore
        load_logistics_cities.cache_clear()  # type: ignore
    elif rows_count > 0:
        # Если были строки в файле, но ни одна не была валидной - это ошибка
        raise ValueError(
            'Не удалось импортировать ни одной записи. '
            'Проверьте, что файл содержит данные и все обязательные поля заполнены.'
        )
    
    return len(cities)


def export_trail_cities_to_excel() -> bytes:
    """Экспортирует города трала в Excel формат.
    
    Returns:
        Байты Excel файла
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError as exc:
        raise RuntimeError('openpyxl не установлен. Установите: pip install openpyxl') from exc
    
    from io import BytesIO
    
    cities = load_trail_cities()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Трал города'
    
    # Стили для заголовков
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # Заголовки
    headers = ['Город', 'Регион', 'Цена трала, руб.']
    ws.append(headers)
    
    # Применяем стили к заголовкам
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Данные
    for city in cities:
        ws.append([
            city.get('name', ''),
            city.get('region', ''),
            city.get('trail_price', 0) or 0
        ])
    
    # Автоподбор ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (AttributeError, TypeError):
                pass  # Игнорируем ошибки преобразования значений ячеек
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в байты
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


def import_trail_cities_from_excel(excel_path: Path, *, actor: Optional[str] = None) -> int:
    """Импортирует города трала из Excel файла, заменяя все существующие данные.
    
    Args:
        excel_path: Путь к Excel файлу
        actor: Имя пользователя, выполняющего импорт (для версионирования)
        
    Returns:
        Количество импортированных записей
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError('openpyxl не установлен. Установите: pip install openpyxl') from exc
    
    if not excel_path.exists():
        raise FileNotFoundError(f'Файл не найден: {excel_path}')
    
    cities = []
    rows_count = 0
    
    wb = None
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Пропускаем заголовок (первая строка)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        rows_count = len(rows)
        
        for row in rows:
            if not row or len(row) < 3:
                continue
            
            # Пропускаем полностью пустые строки
            if not any(cell for cell in row if cell is not None and str(cell).strip()):
                continue
            
            # Индексы колонок: 0-Город, 1-Регион, 2-Цена трала, руб.
            name = str(row[0]).strip() if row[0] else ''
            region = str(row[1]).strip() if row[1] else ''
            trail_price_value = row[2]
            
            # Преобразуем цену трала в число
            try:
                if trail_price_value is None:
                    trail_price = 0.0
                elif isinstance(trail_price_value, (int, float)):
                    trail_price = float(trail_price_value)
                else:
                    trail_price = float(str(trail_price_value).strip().replace(',', '.'))
            except (ValueError, TypeError):
                trail_price = 0.0
            
            if name:
                cities.append({
                    'name': name,
                    'region': region,
                    'trail_price': trail_price
                })
        
    except Exception as exc:
        try:
            current_app.logger.error(f'Ошибка при парсинге Excel: {exc}')
        except RuntimeError:
            pass
        raise
    finally:
        # Гарантированно закрываем файл перед удалением
        if wb is not None:
            wb.close()
    
    # Защита от удаления всех городов: сохраняем только если есть хотя бы одна валидная запись
    if cities:
        save_trail_cities(cities, actor=actor)
        load_trail_cities.cache_clear()  # type: ignore
        load_all_logistics_cities.cache_clear()  # type: ignore
        load_logistics_cities.cache_clear()  # type: ignore
    elif rows_count > 0:
        # Если были строки в файле, но ни одна не была валидной - это ошибка
        raise ValueError(
            'Не удалось импортировать ни одной записи. '
            'Проверьте, что файл содержит данные и все обязательные поля заполнены.'
        )
    
    return len(cities)


@cached_dataset(maxsize=1)
def load_all_logistics_cities() -> List[Dict[str, Any]]:
    """
    Объединяет все три справочника логистики в один список для обратной совместимости.
    Используется в основном приложении для работы с логистикой.
    """
    all_cities = []
    
    # Основные города
    main_cities = load_main_cities()
    for city in main_cities:
        # Проверяем наличие трала в исходных данных
        allows_trail = city.get('allows_trail', False)
        trail_price = city.get('trail_price')
        
        all_cities.append({
            'name': city.get('name', ''),
            'region': city.get('region', ''),
            'truck_price': city.get('truck_price', 0),
            'trail_price': trail_price,
            'is_main_route': city.get('is_main_route', False),
            'allows_trail': allows_trail,
            'distance_from_ekb_km': None,
            **({'main_city': city['main_city']} if 'main_city' in city else {})
        })
    
    # ЕКБ+РФ города
    ekb_rf_cities = load_ekb_rf_cities()
    for city in ekb_rf_cities:
        all_cities.append({
            'name': city.get('name', ''),
            'region': city.get('region', ''),
            'truck_price': 0,  # Для ЕКБ+РФ используется алгоритм, а не прямая цена
            'trail_price': None,
            'is_main_route': False,
            'allows_trail': False,
            'distance_from_ekb_km': city.get('distance_from_ekb_km')
        })
    
    # Трал города
    trail_cities = load_trail_cities()
    for city in trail_cities:
        all_cities.append({
            'name': city.get('name', ''),
            'region': city.get('region', ''),
            'truck_price': 0,  # Для трала используется отдельная цена
            'trail_price': city.get('trail_price', 0),
            'is_main_route': False,
            'allows_trail': True,
            'distance_from_ekb_km': None
        })
    
    return all_cities


def is_city_in_ekb_rf_catalog(city_name: str) -> bool:
    """
    Проверяет, находится ли город в справочнике ЕКБ+РФ.
    
    Args:
        city_name: Название города для проверки
    
    Returns:
        True если город найден в справочнике ЕКБ+РФ, False иначе
    """
    ekb_rf_cities = load_ekb_rf_cities()
    for city in ekb_rf_cities:
        if city.get('name', '').strip() == city_name.strip():
            return True
    return False


def get_ekb_rf_city_distance(city_name: str) -> Optional[int]:
    """
    Получает расстояние от ЕКБ для города из справочника ЕКБ+РФ.
    
    Args:
        city_name: Название города
    
    Returns:
        Расстояние от ЕКБ в километрах или None если город не найден
    """
    ekb_rf_cities = load_ekb_rf_cities()
    for city in ekb_rf_cities:
        if city.get('name', '').strip() == city_name.strip():
            return city.get('distance_from_ekb_km')
    return None


def update_city_distance(city_name: str, distance_km: int, actor: Optional[str] = None) -> bool:
    """
    Обновляет расстояние от ЕКБ для указанного города.
    Ищет город во всех трех справочниках и обновляет в соответствующем.
    
    Args:
        city_name: Название города
        distance_km: Расстояние от Екатеринбурга в километрах
        actor: Кто вносит изменение (username или None)
    
    Returns:
        True если город найден и обновлен, False иначе
    """
    # Ищем в справочнике ЕКБ+РФ (там хранится distance_from_ekb_km)
    ekb_rf_cities = load_ekb_rf_cities()
    for city in ekb_rf_cities:
        if city.get('name') == city_name:
            city['distance_from_ekb_km'] = distance_km
            save_ekb_rf_cities(ekb_rf_cities, actor=actor)
            return True
    
    # Если город не найден в ЕКБ+РФ, но есть расстояние, возможно нужно переместить город
    # или добавить его в справочник ЕКБ+РФ. Для простоты просто возвращаем False.
    return False


@cached_dataset(maxsize=1)
def load_orders_documents() -> List[Dict[str, Any]]:
    """Читает список распоряжений и нормализует структуру файлов."""
    orders_path = CONFIG_DIR / 'orders_documents.json'
    try:
        with orders_path.open('r', encoding='utf-8') as file:
            data = json.load(file)
    except FileNotFoundError:
        _log_error(f'Orders registry file not found at {orders_path.as_posix()}')
        return []
    except json.JSONDecodeError as exc:
        _log_error(f'Failed to parse orders registry file: {exc}')
        return []

    normalized_orders: List[Dict[str, Any]] = []
    for entry in data.get('orders', []):
        files: List[Dict[str, str]] = []
        for file_entry in entry.get('files', []):
            filename = str(file_entry.get('filename', '')).strip()
            if not filename:
                continue
            label = str(
                file_entry.get('label')
                or file_entry.get('format')
                or file_entry.get('name')
                or Path(filename).suffix.replace('.', '').upper()
            ).strip()
            files.append({
                'label': label or 'Скачать',
                'filename': filename
            })

        normalized_orders.append({
            'id': entry.get('id') or entry.get('identifier'),
            'title': entry.get('title') or entry.get('name') or 'Распоряжение',
            'brief': str(entry.get('brief', '')).strip(),
            'summary': str(entry.get('summary', '')).strip(),
            'files': files,
            'updated_at': entry.get('updated_at') or entry.get('date')
        })

    return normalized_orders


def refresh_orders_documents() -> None:
    """Обновляет кэш распоряжений из конфигурационного файла."""
    global ORDERS_REGISTRY
    ORDERS_REGISTRY = load_orders_documents()


def get_orders_documents() -> List[Dict[str, Any]]:
    """Возвращает копию списка распоряжений."""
    return list(ORDERS_REGISTRY)


@cached_dataset(maxsize=1)
def load_task_templates() -> List[Dict[str, Any]]:
    """Читает список шаблонов задач из конфигурационного файла."""
    templates_path = CONFIG_DIR / 'task_templates.json'
    try:
        with templates_path.open('r', encoding='utf-8') as file:
            data = json.load(file)
    except FileNotFoundError:
        _log_error(f'Task templates file not found at {templates_path.as_posix()}')
        return []
    except json.JSONDecodeError as exc:
        _log_error(f'Failed to parse task templates file: {exc}')
        return []

    normalized_templates: List[Dict[str, Any]] = []
    for entry in data.get('templates', []):
        files: List[Dict[str, str]] = []
        for file_entry in entry.get('files', []):
            filename = str(file_entry.get('filename', '')).strip()
            if not filename:
                continue
            label = str(
                file_entry.get('label')
                or file_entry.get('format')
                or file_entry.get('name')
                or Path(filename).suffix.replace('.', '').upper()
            ).strip()
            files.append({
                'label': label or 'Скачать',
                'filename': filename
            })

        normalized_templates.append({
            'id': entry.get('id') or entry.get('identifier'),
            'title': entry.get('title') or entry.get('name') or 'Шаблон',
            'summary': str(entry.get('summary', '')).strip(),
            'files': files,
            'updated_at': entry.get('updated_at') or entry.get('date')
        })

    return normalized_templates


def refresh_task_templates() -> None:
    """Обновляет кэш шаблонов задач."""
    global TASK_TEMPLATES
    TASK_TEMPLATES = load_task_templates()


def get_task_templates() -> List[Dict[str, Any]]:
    """Возвращает копию списка шаблонов задач."""
    return list(TASK_TEMPLATES)


@cached_dataset(maxsize=1)
def load_task_instructions() -> List[Dict[str, Any]]:
    """Читает список инструкций из конфигурационного файла."""
    instructions_path = CONFIG_DIR / 'instructions_tasks.json'
    try:
        with instructions_path.open('r', encoding='utf-8') as file:
            data = json.load(file)
    except FileNotFoundError:
        _log_error(f'Instructions file not found at {instructions_path.as_posix()}')
        return []
    except json.JSONDecodeError as exc:
        _log_error(f'Failed to parse instructions file: {exc}')
        return []

    normalized_instructions: List[Dict[str, Any]] = []
    for entry in data.get('instructions', []):
        files: List[Dict[str, str]] = []
        
        for file_entry in entry.get('files', []):
            filename = str(file_entry.get('filename', '')).strip()
            if not filename:
                continue
            label = str(
                file_entry.get('label')
                or file_entry.get('format')
                or file_entry.get('name')
                or Path(filename).suffix.replace('.', '').upper()
            ).strip()
            files.append({
                'label': label or 'Скачать',
                'filename': filename
            })

        normalized_instructions.append({
            'id': entry.get('id') or entry.get('identifier'),
            'title': entry.get('title') or entry.get('name') or 'Инструкция',
            'brief': str(entry.get('brief', '')).strip(),
            'summary': str(entry.get('summary', '')).strip(),
            'files': files,
            'updated_at': entry.get('updated_at') or entry.get('date')
        })

    return normalized_instructions


def refresh_task_instructions() -> None:
    """Обновляет кэш инструкций."""
    global TASK_INSTRUCTIONS
    TASK_INSTRUCTIONS = load_task_instructions()


def get_task_instructions() -> List[Dict[str, Any]]:
    """
    Возвращает список инструкций с содержимым TXT файлов.
    Содержимое файлов загружается каждый раз при запросе (не кэшируется),
    чтобы изменения в TXT файлах сразу отображались на сайте.
    """
    instructions_dir = BASE_DIR / 'static' / 'instructions'
    result = []
    
    for instruction in TASK_INSTRUCTIONS:
        # Копируем инструкцию
        instruction_copy = dict(instruction)
        
        # Загружаем содержимое TXT файлов каждый раз при запросе
        content_text: str | None = None
        for file_entry in instruction.get('files', []):
            filename = str(file_entry.get('filename', '')).strip()
            if filename.lower().endswith('.txt'):
                file_path = instructions_dir / filename
                if file_path.exists():
                    try:
                        with file_path.open('r', encoding='utf-8') as txt_file:
                            content_text = txt_file.read()
                            break  # Берем первый найденный TXT файл
                    except Exception as exc:
                        _log_error(f'Failed to read instruction file {filename}: {exc}')
        
        instruction_copy['content_text'] = content_text
        result.append(instruction_copy)
    
    return result


def parse_steel_prices_csv(csv_path: Path) -> List[Dict[str, Any]]:
    """Парсит CSV файл со сталями и ценами, извлекая данные о материалах.
    
    Args:
        csv_path: Путь к CSV файлу
        
    Returns:
        Список словарей с полями: russian, gb, gost, price, workpiece_type, notes
    """
    materials = []
    
    try:
        with csv_path.open('r', encoding='utf-8') as file:
            reader = csv.reader(file)
            rows = list(reader)
            
            # Пропускаем служебные строки, ищем заголовки (строка 9, индекс 8)
            # Данные начинаются со строки 10 (индекс 9)
            for i, row in enumerate(rows):
                if i < 9:  # Пропускаем служебные строки до заголовков
                    continue
                
                # Пропускаем пустые строки или строки без данных
                if len(row) < 9:
                    continue
                
                # Извлекаем данные из колонок
                first_col = row[0].strip() if len(row) > 0 else ''
                russian = row[1].strip() if len(row) > 1 else ''
                gb = row[4].strip() if len(row) > 4 else ''
                gost = row[6].strip() if len(row) > 6 else ''
                price = row[7].strip() if len(row) > 7 else ''
                workpiece_type = row[8].strip() if len(row) > 8 else ''
                
                # Пропускаем заголовки (строка с "№ в группе" или "Материал" в первой колонке)
                if first_col in ('№ в группе', 'Материал') or russian in ('Материал', 'Наименование мир'):
                    continue
                
                # Пропускаем группирующие строки (начинаются с пробела и цифры с точкой, например " 1. Углеродистая")
                if first_col and len(first_col) > 2 and first_col[0] == ' ':
                    if first_col[1].isdigit() and '.' in first_col[:5]:
                        continue
                
                # Пропускаем строки без российского материала
                if not russian:
                    continue
                
                # Пропускаем группирующие строки (если российский материал начинается с пробела, цифры и точки)
                # и содержит длинный текст (например " 1. Углеродистая (нелегир)")
                if russian and len(russian) > 15 and russian[0] == ' ':
                    if russian[1].isdigit() and '.' in russian[:5]:
                        continue
                
                materials.append({
                    'russian': russian,
                    'gb': gb,
                    'gost': gost,
                    'price': price,
                    'workpiece_type': workpiece_type,
                    'notes': ''
                })
                
    except FileNotFoundError:
        try:
            _log_error(f'CSV file not found at {csv_path.as_posix()}')
        except RuntimeError:
            pass  # Игнорируем ошибку, если нет application context
    except Exception as exc:
        try:
            _log_error(f'Failed to parse CSV file: {exc}')
        except RuntimeError:
            pass  # Игнорируем ошибку, если нет application context
    
    return materials


def import_gb_materials_from_csv(csv_path: Path, *, actor: Optional[str] = None) -> int:
    """Импортирует материалы из CSV файла, заменяя все существующие данные.
    
    Args:
        csv_path: Путь к CSV файлу
        actor: Имя пользователя, выполняющего импорт (для версионирования)
        
    Returns:
        Количество импортированных записей
    """
    materials = parse_steel_prices_csv(csv_path)
    
    if materials:
        save_gb_materials(materials, actor=actor)
        refresh_gb_analogs()
    
    return len(materials)


def import_gb_materials_from_excel(excel_path: Path, *, actor: Optional[str] = None) -> int:
    """Импортирует материалы из Excel файла, заменяя все существующие данные.
    
    Args:
        excel_path: Путь к Excel файлу
        actor: Имя пользователя, выполняющего импорт (для версионирования)
        
    Returns:
        Количество импортированных записей
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError('openpyxl не установлен. Установите: pip install openpyxl') from exc
    
    if not excel_path.exists():
        raise FileNotFoundError(f'Файл не найден: {excel_path}')
    
    materials = []
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Пропускаем заголовок (первая строка)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        for row in rows:
            if not row or len(row) < 9:
                continue
            
            # Индексы колонок: 0-номер, 1-материал(RU), 2-4 пустые, 5-GB, 6 пустая, 7-ГОСТ, 8-цена, 9-вид заготовки
            russian = str(row[1]).strip() if row[1] else ''
            gb = str(row[4]).strip() if row[4] else ''
            gost = str(row[6]).strip() if row[6] else ''
            price = str(row[7]).strip() if row[7] else ''
            workpiece_type = str(row[8]).strip() if row[8] else ''
            
            if russian and gb:
                materials.append({
                    'russian': russian,
                    'gb': gb,
                    'notes': '',
                    'gost': gost,
                    'price': price,
                    'workpiece_type': workpiece_type
                })
        
        wb.close()
        
    except Exception as exc:
        try:
            current_app.logger.error(f'Ошибка при парсинге Excel: {exc}')
        except RuntimeError:
            pass
        raise
    
    if materials:
        save_gb_materials(materials, actor=actor)
        refresh_gb_analogs()
    
    return len(materials)


def export_gb_materials_to_excel() -> bytes:
    """Экспортирует материалы GB в Excel формат.
    
    Returns:
        Байты Excel файла
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError as exc:
        raise RuntimeError('openpyxl не установлен. Установите: pip install openpyxl') from exc
    
    from io import BytesIO
    
    materials = load_gb_materials()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Материалы'
    
    # Стили для заголовков
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # Заголовки
    headers = ['№ в группе', 'Материал', '', '', 'Наименование мир', '', 'ГОСТ', 'Материал.Цена', 'Материал.Цена.Вид заготовки']
    ws.append(headers)
    
    # Применяем стили к заголовкам
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Данные
    for i, material in enumerate(materials, start=1):
        ws.append([
            i,
            material.get('russian', ''),
            '',
            '',
            material.get('gb', ''),
            '',
            material.get('gost', ''),
            material.get('price', ''),
            material.get('workpiece_type', '')
        ])
    
    # Автоподбор ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (AttributeError, TypeError):
                pass  # Игнорируем ошибки преобразования значений ячеек
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в байты
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


def export_duty_rates_to_excel() -> bytes:
    """Экспортирует простые ставки пошлин (без кода ТН ВЭД) в Excel формат.
    
    Returns:
        Байты Excel файла
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError as exc:
        raise RuntimeError('openpyxl не установлен. Установите: pip install openpyxl') from exc
    
    from io import BytesIO
    
    # Загружаем все пошлины и фильтруем только простые (без code)
    all_items = load_duty_rates()
    duty_items = [item for item in all_items if not item.get('code')]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Пошлины'
    
    # Стили для заголовков
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # Заголовки
    headers = ['Товар', 'Категория', 'Пошлина, %']
    ws.append(headers)
    
    # Применяем стили к заголовкам
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Данные
    for item in duty_items:
        ws.append([
            item.get('product', ''),
            item.get('category', ''),
            item.get('duty_percent', 0.0)
        ])
    
    # Автоподбор ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (AttributeError, TypeError):
                pass  # Игнорируем ошибки преобразования значений ячеек
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в байты
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


def import_duty_rates_from_excel(excel_path: Path, *, actor: Optional[str] = None) -> int:
    """Импортирует ставки пошлин из Excel файла, заменяя все существующие данные.
    
    Args:
        excel_path: Путь к Excel файлу
        actor: Имя пользователя, выполняющего импорт (для версионирования)
        
    Returns:
        Количество импортированных записей
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError('openpyxl не установлен. Установите: pip install openpyxl') from exc
    
    if not excel_path.exists():
        raise FileNotFoundError(f'Файл не найден: {excel_path}')
    
    duty_items = []
    
    wb = None
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Пропускаем заголовок (первая строка)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        for row in rows:
            if not row or len(row) < 3:
                continue
            
            # Индексы колонок: 0-Товар, 1-Категория, 2-Пошлина, %
            product = str(row[0]).strip() if row[0] else ''
            category = str(row[1]).strip() if row[1] else ''
            duty_percent_value = row[2]
            
            # Преобразуем пошлину в число
            try:
                if duty_percent_value is None:
                    duty_percent = 0.0
                elif isinstance(duty_percent_value, (int, float)):
                    duty_percent = float(duty_percent_value)
                else:
                    duty_percent = float(str(duty_percent_value).strip().replace(',', '.'))
            except (ValueError, TypeError):
                duty_percent = 0.0
            
            if product and category:
                duty_items.append({
                    'product': product,
                    'category': category,
                    'duty_percent': duty_percent
                })
        
    except Exception as exc:
        try:
            current_app.logger.error(f'Ошибка при парсинге Excel: {exc}')
        except RuntimeError:
            pass
        raise
    finally:
        # Гарантированно закрываем файл перед удалением
        if wb is not None:
            wb.close()
    
    if duty_items:
        # Сохраняем в единый файл tnved_catalog.json
        # Загружаем существующие записи и добавляем новые
        all_items = load_duty_rates()
        # Заменяем только простые записи (без code), сохраняем ТН ВЭД
        tnved_items = [item for item in all_items if item.get('code')]
        all_items = duty_items + tnved_items
        save_duty_rates(all_items, actor=actor)
        refresh_duty_rates()
    
    return len(duty_items)


def export_tnved_catalog_to_excel() -> bytes:
    """Экспортирует каталог ТН ВЭД в Excel формат.
    
    Экспортирует только записи с полем 'code' (ТН ВЭД) из duty_rates.json.
    
    Returns:
        Байты Excel файла
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError as exc:
        raise RuntimeError('openpyxl не установлен. Установите: pip install openpyxl') from exc
    
    from io import BytesIO
    
    # Загружаем все пошлины и фильтруем только ТН ВЭД (с полем code)
    all_items = load_duty_rates()
    tnved_items = [item for item in all_items if item.get('code')]
    wb = Workbook()
    ws = wb.active
    ws.title = 'ТН ВЭД'
    
    # Стили для заголовков
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # Заголовки
    headers = ['Код ТН ВЭД', 'Описание', 'Ключевые слова', 'Примеры', 'Пошлина', 'Пошлина, %']
    ws.append(headers)
    
    # Применяем стили к заголовкам
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Данные
    for item in tnved_items:
        ws.append([
            item.get('code', ''),
            item.get('description', ''),
            item.get('keywords_display', ''),
            item.get('examples', ''),
            item.get('duty_text', ''),
            item.get('duty_percent') if item.get('duty_percent') is not None else ''
        ])
    
    # Автоподбор ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (AttributeError, TypeError):
                pass  # Игнорируем ошибки преобразования значений ячеек
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в байты
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


def import_tnved_catalog_from_excel(excel_path: Path, *, actor: Optional[str] = None) -> int:
    """Импортирует каталог ТН ВЭД из Excel файла, заменяя все существующие данные.
    
    Args:
        excel_path: Путь к Excel файлу
        actor: Имя пользователя, выполняющего импорт (для версионирования)
        
    Returns:
        Количество импортированных записей
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError('openpyxl не установлен. Установите: pip install openpyxl') from exc
    
    if not excel_path.exists():
        raise FileNotFoundError(f'Файл не найден: {excel_path}')
    
    tnved_items = []
    
    wb = None
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Пропускаем заголовок (первая строка)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        for row in rows:
            if not row or len(row) < 5:
                continue
            
            # Индексы колонок: 0-Код ТН ВЭД, 1-Описание, 2-Ключевые слова, 3-Примеры, 4-Пошлина, 5-Пошлина, %
            code = str(row[0]).strip() if row[0] else ''
            description = str(row[1]).strip() if row[1] else ''
            keywords_text = str(row[2]).strip() if row[2] else ''
            examples_text = str(row[3]).strip() if row[3] else ''
            duty_text = str(row[4]).strip() if row[4] else ''
            duty_percent_value = row[5] if len(row) > 5 else None
            
            # Преобразуем пошлину в число
            duty_percent = None
            if duty_percent_value is not None:
                try:
                    if isinstance(duty_percent_value, (int, float)):
                        duty_percent = float(duty_percent_value)
                    else:
                        duty_percent = float(str(duty_percent_value).strip().replace(',', '.'))
                except (ValueError, TypeError):
                    # Если не удалось преобразовать, пытаемся извлечь из duty_text
                    duty_percent = _extract_percent_value(duty_text)
            else:
                # Если не указано, пытаемся извлечь из duty_text
                duty_percent = _extract_percent_value(duty_text)
            
            if code:
                tnved_items.append({
                    'code': code,
                    'description': description,
                    'keywords_display': keywords_text,
                    'examples': examples_text,
                    'duty_text': duty_text,
                    'duty_percent': duty_percent
                })
        
    except Exception as exc:
        try:
            current_app.logger.error(f'Ошибка при парсинге Excel: {exc}')
        except RuntimeError:
            pass
        raise
    finally:
        # Гарантированно закрываем файл перед удалением
        if wb is not None:
            wb.close()
    
    if tnved_items:
        # Сохраняем в единый файл tnved_catalog.json
        save_duty_rates(tnved_items, actor=actor)
        refresh_duty_rates()
    
    return len(tnved_items)


def save_orders_documents(orders: List[Dict[str, Any]], *, actor: Optional[str] = None) -> None:
    """Сохраняет список распоряжений и обновляет кэш."""
    orders_path = CONFIG_DIR / 'orders_documents.json'
    orders_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {'orders': orders}
    _snapshot_version('orders_documents', payload, actor)

    with orders_path.open('w', encoding='utf-8') as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)

    refresh_orders_documents()
    # Инвалидируем кэш после сохранения
    load_orders_documents.cache_clear()  # type: ignore


def save_task_templates(templates: List[Dict[str, Any]], *, actor: Optional[str] = None) -> None:
    """Сохраняет перечень шаблонов и обновляет кэш."""
    templates_path = CONFIG_DIR / 'task_templates.json'
    templates_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {'templates': templates}
    _snapshot_version('task_templates', payload, actor)

    with templates_path.open('w', encoding='utf-8') as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)

    refresh_task_templates()
    # Инвалидируем кэш после сохранения
    load_task_templates.cache_clear()  # type: ignore


def save_task_instructions(instructions: List[Dict[str, Any]], *, actor: Optional[str] = None) -> None:
    """Сохраняет список инструкций и обновляет кэш."""
    instructions_path = CONFIG_DIR / 'instructions_tasks.json'
    instructions_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {'instructions': instructions}
    _snapshot_version('instructions_tasks', payload, actor)

    with instructions_path.open('w', encoding='utf-8') as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)

    refresh_task_instructions()
    # Инвалидируем кэш после сохранения
    load_task_instructions.cache_clear()  # type: ignore


def save_with_version(collection: str, data: Dict[str, Any], *, actor: Optional[str] = None) -> None:
    """Обёртка для сохранения произвольных данных с версионированием."""
    _snapshot_version(collection, data, actor)


def parse_files_input(raw_text: str) -> List[Dict[str, str]]:
    """Преобразует текстовое представление файлов контента в структуру."""
    if not raw_text:
        return []

    files: List[Dict[str, str]] = []
    for line in raw_text.splitlines():
        cleaned = line.strip()
        if not cleaned:
            continue

        if '|' in cleaned:
            label, filename = cleaned.split('|', 1)
        elif ';' in cleaned:
            label, filename = cleaned.split(';', 1)
        else:
            parts = cleaned.split(maxsplit=1)
            label = parts[0]
            filename = parts[1] if len(parts) > 1 else parts[0]

        files.append({'label': label.strip(), 'filename': filename.strip()})

    return files


def format_files_output(files: List[Dict[str, Any]]) -> str:
    """Формирует читабельное представление файлов для textarea."""
    lines = []
    for file_item in files:
        label = str(file_item.get('label', '')).strip()
        filename = str(file_item.get('filename', '')).strip()
        if label and filename:
            lines.append(f'{label} | {filename}')
        elif filename:
            lines.append(filename)
    return '\n'.join(lines)


def make_content_id(prefix: str) -> str:
    """Создаёт уникальный идентификатор контента."""
    return f'{prefix}-{uuid4().hex[:8]}'


def list_content_versions(collection: str) -> List[Dict[str, Any]]:
    """Возвращает список версий для указанной коллекции."""
    versions_path = VERSIONS_DIR / collection
    if not versions_path.exists():
        return []

    versions: List[Dict[str, Any]] = []
    for entry in sorted(versions_path.iterdir(), reverse=True):
        if entry.suffix != '.json' or not entry.is_file():
            continue
        try:
            with entry.open('r', encoding='utf-8') as file:
                payload = json.load(file)
        except (OSError, json.JSONDecodeError):
            continue

        versions.append({
            'filename': entry.name,
            'saved_at': payload.get('saved_at'),
            'actor': payload.get('actor'),
        })

    return versions


def load_content_version(collection: str, filename: str) -> Optional[Dict[str, Any]]:
    """Загружает выбранную версию контента."""
    target = VERSIONS_DIR / collection / filename
    if not target.exists():
        return None
    try:
        with target.open('r', encoding='utf-8') as file:
            return json.load(file)
    except (OSError, json.JSONDecodeError):
        return None


def init_app(_app: Any) -> None:
    """Инициализирует кэшированные данные при старте приложения."""
    refresh_gb_analogs()
    refresh_duty_rates()
    refresh_tnved_catalog()
    refresh_orders_documents()
    refresh_task_templates()
    refresh_task_instructions()


def _snapshot_version(collection: str, payload: Dict[str, Any], actor: Optional[str] = None) -> None:
    """Сохраняет резервную копию данных в каталоге версий."""
    if not payload:
        return

    timestamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%S')
    versions_path = VERSIONS_DIR / collection
    versions_path.mkdir(parents=True, exist_ok=True)

    filename = f'{timestamp}.json'
    meta = {
        'saved_at': timestamp,
        'actor': actor,
        'data': payload,
    }

    try:
        with (versions_path / filename).open('w', encoding='utf-8') as file:
            json.dump(meta, file, ensure_ascii=False, indent=2)
    except OSError:
        _log_error(f'Failed to persist version snapshot for {collection}')
