from __future__ import annotations

import logging
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import BinaryIO, Dict, Iterable, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException

from app.business.interfaces import ExcelImporterPort
from app.core.exceptions import ImportError as KPImportError


HeaderMap = Dict[str, str]


@dataclass(frozen=True)
class ImportSchema:
    header_map: HeaderMap
    required_fields: Iterable[str]
    numeric_fields: Iterable[str]

    def normalise_header(self, value: object) -> str:
        if value is None:
            return ''
        return str(value).strip().lower()

    def resolve_field(self, header_value: object) -> str | None:
        normalised = self.normalise_header(header_value)
        return self.header_map.get(normalised)


SCHEMA = ImportSchema(
    header_map={
        '№': 'ordinal',
        'no': 'ordinal',
        '#': 'ordinal',
        'номер позиции': 'ordinal',
        'номенклатура': 'product',
        'наименование товара': 'product',
        'наименование продукции': 'product',
        'название': 'product',
        'номер чертежа': 'drawing_number',
        'чертеж': 'drawing_number',
        'материал': 'material',
        'материал изделия': 'material',
        'цена закупа': 'cost_price',
        'цена закупа, руб': 'cost_price',
        'стоимость закупа': 'cost_price',
        'цена закупа за кг': 'cost_price_per_kg',
        'цена за кг': 'cost_price_per_kg',
        'цена закупа за кг, руб': 'cost_price_per_kg',
        'стоимость закупа за кг': 'cost_price_per_kg',
        'количество, шт.': 'quantity',
        'количество шт.': 'quantity',
        'количество': 'quantity',
        'шт.': 'quantity',
        'вес за шт. (кг)': 'weight',
        'вес за шт. кг': 'weight',
        'вес, кг': 'weight',
        'вес': 'weight',
        'пошлина (%)': 'duty_percent',
        'пошлина %': 'duty_percent',
        'пошлина': 'duty_percent',
    },
    required_fields=('product', 'quantity', 'weight'),
    numeric_fields=('cost_price', 'cost_price_per_kg', 'quantity', 'weight', 'duty_percent'),
)


FIELD_LABELS = {
    'product': 'Номенклатура',
    'drawing_number': 'Номер чертежа',
    'material': 'Материал',
    'cost_price': 'Цена закупа',
    'cost_price_per_kg': 'Цена закупа за кг',
    'quantity': 'Количество, шт.',
    'weight': 'Вес за шт. (кг)',
    'duty_percent': 'Пошлина (%)',
    'ordinal': '№',
}


class ExcelImportError(ValueError):
    """Ошибка валидации Excel при импорте позиций."""


def _normalise_numeric(field: str, value: object, row_idx: int) -> str:
    if value is None:
        return ''

    if isinstance(value, (int, float, Decimal)):
        decimal_value = Decimal(str(value))
    elif isinstance(value, str):
        normalised = value.replace(' ', '').replace(',', '.').strip()
        if not normalised:
            return ''
        try:
            decimal_value = Decimal(normalised)
        except InvalidOperation as exc:
            raise ExcelImportError(
                f"Некорректное числовое значение в столбце '{FIELD_LABELS[field]}' (строка {row_idx})."
            ) from exc
    else:
        raise ExcelImportError(
            f"Некорректный формат данных в столбце '{FIELD_LABELS[field]}' (строка {row_idx})."
        )

    if field == 'quantity':
        if decimal_value != decimal_value.to_integral_value():
            raise ExcelImportError(
                f"Количество в строке {row_idx} должно быть целым числом."
            )
        return str(int(decimal_value))

    normalised_dec = decimal_value.normalize()
    if normalised_dec == normalised_dec.to_integral_value():
        return str(int(normalised_dec))
    return format(normalised_dec, 'f')


def _prepare_value(field: str, value: object, row_idx: int) -> str:
    if field in SCHEMA.numeric_fields:
        return _normalise_numeric(field, value, row_idx)

    if value is None:
        return ''
    return str(value).strip()


def _validate_and_calculate_prices(
    cost_price: str,
    cost_price_per_kg: str,
    weight: str,
    quantity: str,
    row_idx: int
) -> tuple[str, str]:
    """
    Валидирует и вычисляет цены закупа.
    
    Возвращает кортеж (cost_price, cost_price_per_kg) с валидными значениями.
    Выбрасывает ExcelImportError при несоответствии цен.
    """
    from decimal import Decimal, InvalidOperation
    
    # Преобразуем в Decimal для точных вычислений
    def to_decimal(value: str) -> Decimal | None:
        if not value or value.strip() == '':
            return None
        try:
            return Decimal(str(value).replace(',', '.').strip())
        except (InvalidOperation, ValueError):
            return None
    
    cost_price_dec = to_decimal(cost_price)
    cost_price_per_kg_dec = to_decimal(cost_price_per_kg)
    weight_dec = to_decimal(weight)
    quantity_dec = to_decimal(quantity)
    
    # Проверяем, что указана хотя бы одна цена
    if cost_price_dec is None and cost_price_per_kg_dec is None:
        raise ExcelImportError(
            f"В строке {row_idx} должна быть указана либо 'Цена закупа', либо 'Цена закупа за кг'."
        )
    
    # Проверяем наличие веса и количества для расчетов
    if weight_dec is None or weight_dec <= 0:
        raise ExcelImportError(
            f"В строке {row_idx} должен быть указан вес для расчета цен."
        )
    
    if quantity_dec is None or quantity_dec <= 0:
        raise ExcelImportError(
            f"В строке {row_idx} должно быть указано количество для расчета цен."
        )
    
    # cost_price - это цена за единицу продукции (не за все количество!)
    # cost_price_per_kg - это цена за кг
    # Формула: cost_price = cost_price_per_kg * weight (цена за единицу = цена за кг * вес единицы)
    
    # Если указаны обе цены, проверяем их согласованность
    if cost_price_dec is not None and cost_price_per_kg_dec is not None:
        # Вычисляем ожидаемую цену за единицу из цены за кг
        expected_cost_price = cost_price_per_kg_dec * weight_dec
        
        # Допустимое отклонение: 1% или минимум 0.01 руб
        tolerance = max(expected_cost_price * Decimal('0.01'), Decimal('0.01'))
        difference = abs(cost_price_dec - expected_cost_price)
        
        if difference > tolerance:
            # Форматируем числа для читаемости
            def format_decimal(d: Decimal, decimals: int = 2) -> str:
                return f"{float(d):.{decimals}f}".rstrip('0').rstrip('.')
            
            raise ExcelImportError(
                f"В строке {row_idx} обнаружено несоответствие цен:\n"
                f"  - Цена закупа за кг: {format_decimal(cost_price_per_kg_dec)}/кг\n"
                f"  - Вес за шт.: {format_decimal(weight_dec, 3)} кг\n"
                f"  - Количество: {int(quantity_dec)} шт.\n"
                f"  - Ожидаемая цена закупа за единицу: {format_decimal(expected_cost_price)}\n"
                f"  - Указанная цена закупа за единицу: {format_decimal(cost_price_dec)}\n"
                f"  - Разница: {format_decimal(difference)}\n"
                f"Проверьте правильность указанных значений."
            )
        # Если цены согласованы, используем их как есть
        return str(cost_price_dec), str(cost_price_per_kg_dec)
    
    # Если указана только цена за кг, вычисляем цену за единицу
    if cost_price_per_kg_dec is not None:
        calculated_cost_price = cost_price_per_kg_dec * weight_dec
        return str(calculated_cost_price), str(cost_price_per_kg_dec)
    
    # Если указана только цена за единицу, вычисляем цену за кг
    if cost_price_dec is not None:
        calculated_cost_price_per_kg = cost_price_dec / weight_dec if weight_dec > 0 else Decimal('0')
        return str(cost_price_dec), str(calculated_cost_price_per_kg)
    
    # Этот случай не должен произойти, но на всякий случай
    raise ExcelImportError(f"В строке {row_idx} не удалось определить цены.")


class ExcelImporterService(ExcelImporterPort):
    """Стандартная реализация импорта позиций из Excel."""

    def parse_positions(self, stream: BinaryIO) -> List[Dict[str, str]]:
        """Читает Excel шаблон и возвращает список позиций для заполнения формы."""

        if hasattr(stream, 'seek'):
            stream.seek(0)

        try:
            workbook = load_workbook(stream, data_only=True)
        except InvalidFileException as exc:
            raise ExcelImportError('Не удалось прочитать файл Excel. Убедитесь, что используется формат .xlsx.') from exc
        except (IOError, OSError) as exc:
            raise KPImportError(
                'Ошибка чтения файла. Проверьте, что файл не поврежден и доступен для чтения.',
                file_type='xlsx'
            ) from exc
        except Exception as exc:  # pragma: no cover - делегирование неожиданных ошибок
            raise ExcelImportError('Файл повреждён или имеет неподдерживаемый формат.') from exc

        sheet = workbook.active

        # Сначала пытаемся распарсить как старый формат (по заголовкам)
        try:
            return self._parse_with_headers(sheet)
        except ExcelImportError as header_exc:
            # Если не получилось (нет заголовков или обязательных колонок), пробуем новый формат
            # Проверяем, что ошибка связана именно с отсутствием заголовков/колонок
            error_msg = str(header_exc).lower()
            if (
                'заголовков' in error_msg
                or 'обязательные столбцы' in error_msg
                or 'отсутствуют обязательные столбцы' in error_msg
            ):
                try:
                    return self._parse_fixed_template(sheet)
                except ExcelImportError as fixed_exc:
                    # Если и новый формат не подошел, сообщаем подробную информацию по обоим вариантам
                    detailed_message = (
                        "Не удалось импортировать файл ни как таблицу с заголовками, "
                        "ни как шаблон запроса с фиксированными ячейками.\n\n"
                        f"Попытка 1 (формат с заголовками): {header_exc}\n"
                        f"Попытка 2 (формат шаблона запроса): {fixed_exc}"
                    )
                    raise ExcelImportError(detailed_message) from fixed_exc
            # Если ошибка не связана с форматом, пробрасываем её как есть
            raise header_exc

    def _parse_with_headers(self, sheet: Worksheet) -> List[Dict[str, str]]:
        """Парсит Excel файл со старой схемой (по заголовкам в первой строке)."""

        header_row = None
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            if any(str(cell.value).strip() for cell in row if cell.value is not None):
                header_row = row
                break

        if not header_row:
            raise ExcelImportError('Файл не содержит заголовков и не может быть импортирован.')

        column_map: Dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            field_name = SCHEMA.resolve_field(cell.value)
            if field_name and field_name not in column_map:
                column_map[field_name] = idx

        missing = [FIELD_LABELS[field] for field in SCHEMA.required_fields if field not in column_map]
        if missing:
            missing_list = ', '.join(missing)
            raise ExcelImportError(f'В файле отсутствуют обязательные столбцы: {missing_list}.')

        positions: List[Dict[str, str]] = []
        start_row = header_row[0].row + 1

        for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row):
            row_idx = row[0].row
            row_payload: Dict[str, str] = {}

            for field, column_index in column_map.items():
                value = row[column_index].value if column_index < len(row) else None
                row_payload[field] = _prepare_value(field, value, row_idx)

            if not any(row_payload.get(field) for field in SCHEMA.required_fields):
                continue

            for field in SCHEMA.required_fields:
                if not row_payload.get(field):
                    raise ExcelImportError(
                        f"В строке {row_idx} отсутствует значение в столбце '{FIELD_LABELS[field]}'."
                    )

            # Валидация и расчет цен
            cost_price = row_payload.get('cost_price', '')
            cost_price_per_kg = row_payload.get('cost_price_per_kg', '')
            weight = row_payload.get('weight', '')
            quantity = row_payload.get('quantity', '')

            # Валидируем и вычисляем цены
            validated_cost_price, validated_cost_price_per_kg = _validate_and_calculate_prices(
                cost_price, cost_price_per_kg, weight, quantity, row_idx
            )

            # Сохраняем вычисленные значения
            row_payload['cost_price'] = validated_cost_price
            row_payload['cost_price_per_kg'] = validated_cost_price_per_kg

            if 'duty_percent' in row_payload:
                row_payload['duty_percent'] = row_payload.get('duty_percent') or '0'
            else:
                row_payload['duty_percent'] = '0'
            positions.append(row_payload)

        if not positions:
            raise ExcelImportError('Файл не содержит позиций для импорта.')

        return positions

    def _parse_fixed_template(self, sheet: Worksheet) -> List[Dict[str, str]]:
        """
        Парсит Excel файл с новой схемой (фиксированные ячейки начиная с 7-й строки).

        Маппинг:
        - B (колонка 2) -> product (Наименование)
        - D (колонка 4) -> material (Материал)
        - E (колонка 5) -> drawing_number (Номер чертежа)
        - K (колонка 11) -> quantity (Количество)
        - L (колонка 12) -> cost_price (Цена закупа за шт.)
        - M (колонка 13) -> total_cost (общая цена закупа, только для валидации)
        - N (колонка 14) -> weight (Вес за шт.)
        - P (колонка 16) -> duty_percent (Пошлина, %)
        - S (колонка 19) -> cost_price_per_kg (Цена закупа за кг)
        """

        positions: List[Dict[str, str]] = []
        start_row = 7
        max_row = sheet.max_row

        for row_idx in range(start_row, max_row + 1):
            # Читаем значения из фиксированных ячеек
            # openpyxl использует 1-based индексацию для строк и колонок
            product_cell = sheet.cell(row=row_idx, column=2)  # B
            material_cell = sheet.cell(row=row_idx, column=4)  # D
            drawing_number_cell = sheet.cell(row=row_idx, column=5)  # E
            quantity_cell = sheet.cell(row=row_idx, column=11)  # K
            cost_price_cell = sheet.cell(row=row_idx, column=12)  # L
            total_cost_cell = sheet.cell(row=row_idx, column=13)  # M (для валидации)
            weight_cell = sheet.cell(row=row_idx, column=14)  # N
            duty_percent_cell = sheet.cell(row=row_idx, column=16)  # P
            cost_price_per_kg_cell = sheet.cell(row=row_idx, column=19)  # S

            # Формируем row_payload
            row_payload: Dict[str, str] = {}

            # Обязательные поля
            row_payload['product'] = _prepare_value('product', product_cell.value, row_idx)
            row_payload['quantity'] = _prepare_value('quantity', quantity_cell.value, row_idx)
            row_payload['weight'] = _prepare_value('weight', weight_cell.value, row_idx)

            # Строки без количества считаем не позициями:
            # - до первой считанной позиции просто пропускаем их (могут быть пустые/служебные строки),
            # - после того как уже были считаны позиции — считаем концом списка и прекращаем чтение.
            if not row_payload['quantity']:
                if positions:
                    break
                continue

            # Опциональные поля
            row_payload['drawing_number'] = _prepare_value('drawing_number', drawing_number_cell.value, row_idx)
            row_payload['material'] = _prepare_value('material', material_cell.value, row_idx)
            row_payload['cost_price'] = _prepare_value('cost_price', cost_price_cell.value, row_idx)
            # Цена за кг может быть указана отдельно в колонке S
            row_payload['cost_price_per_kg'] = _prepare_value(
                'cost_price_per_kg',
                cost_price_per_kg_cell.value,
                row_idx,
            )
            row_payload['duty_percent'] = _prepare_value('duty_percent', duty_percent_cell.value, row_idx) or '0'

            # Проверяем обязательные поля
            for field in SCHEMA.required_fields:
                if not row_payload.get(field):
                    raise ExcelImportError(
                        f"В строке {row_idx} отсутствует значение в столбце '{FIELD_LABELS[field]}'."
                    )

            # Валидация общей суммы (M) - мягкая проверка с логированием
            total_cost_value = total_cost_cell.value
            if total_cost_value is not None:
                try:
                    total_cost_dec = Decimal(str(total_cost_value).replace(',', '.').strip())
                    cost_price_dec = Decimal(row_payload['cost_price']) if row_payload['cost_price'] else None
                    quantity_dec = Decimal(row_payload['quantity']) if row_payload['quantity'] else None

                    if cost_price_dec and quantity_dec:
                        expected_total = cost_price_dec * quantity_dec
                        tolerance = max(expected_total * Decimal('0.01'), Decimal('0.01'))
                        difference = abs(total_cost_dec - expected_total)

                        if difference > tolerance:
                            # Логируем предупреждение, но не падаем
                            logger = logging.getLogger(__name__)
                            logger.warning(
                                f"В строке {row_idx} обнаружено несоответствие общей суммы закупа:\n"
                                f"  - Указанная общая сумма: {total_cost_dec}\n"
                                f"  - Ожидаемая (цена × количество): {expected_total}\n"
                                f"  - Разница: {difference}"
                            )
                except (ValueError, InvalidOperation):
                    # Игнорируем ошибки парсинга общей суммы
                    pass

            # Валидация и расчет цен
            cost_price = row_payload.get('cost_price', '')
            cost_price_per_kg = row_payload.get('cost_price_per_kg', '')
            weight = row_payload.get('weight', '')
            quantity = row_payload.get('quantity', '')

            # Валидируем и вычисляем цены
            validated_cost_price, validated_cost_price_per_kg = _validate_and_calculate_prices(
                cost_price, cost_price_per_kg, weight, quantity, row_idx
            )

            # Сохраняем вычисленные значения
            row_payload['cost_price'] = validated_cost_price
            row_payload['cost_price_per_kg'] = validated_cost_price_per_kg

            positions.append(row_payload)

        if not positions:
            raise ExcelImportError('Файл не содержит позиций для импорта.')

        return positions


_default_importer = ExcelImporterService()


def parse_positions_from_excel(stream: BinaryIO) -> List[Dict[str, str]]:
    return _default_importer.parse_positions(stream)


__all__ = ['ExcelImportError', 'ExcelImporterService', 'parse_positions_from_excel']
