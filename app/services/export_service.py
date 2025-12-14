"""Сервис для экспорта данных в Excel и PDF."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional

from flask import Response

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def export_audit_logs_to_excel(logs_data: Dict[str, Any]) -> BytesIO:
    """
    Экспортирует логи аудита в Excel.

    Args:
        logs_data: Словарь с данными логов (из AuditLogRepository.get_logs)

    Returns:
        BytesIO буфер с Excel файлом
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError('openpyxl не установлен. Установите: pip install openpyxl')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Лог действий'

    # Заголовки
    headers = ['ID', 'Время', 'Пользователь', 'Действие', 'Ресурс', 'ID ресурса', 'Описание', 'IP-адрес']
    ws.append(headers)

    # Стили для заголовков
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Данные
    for log in logs_data.get('items', []):
        row = [
            log.get('id'),
            log.get('created_at', ''),
            log.get('username', ''),
            log.get('action_type', ''),
            log.get('resource_type', ''),
            log.get('resource_id', ''),
            log.get('description', ''),
            log.get('ip_address', ''),
        ]
        ws.append(row)

    # Автоподбор ширины колонок
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Сохранение в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_audit_logs_to_pdf(logs_data: Dict[str, Any]) -> BytesIO:
    """
    Экспортирует логи аудита в PDF.

    Args:
        logs_data: Словарь с данными логов (из AuditLogRepository.get_logs)

    Returns:
        BytesIO буфер с PDF файлом
    """
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError('reportlab не установлен. Установите: pip install reportlab')

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []

    # Стили
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
    )
    normal_style = styles['Normal']

    # Заголовок
    title = Paragraph('Лог действий пользователей', title_style)
    story.append(title)
    story.append(Spacer(1, 12))

    # Информация о периоде
    info_text = f"Всего записей: {logs_data.get('pagination', {}).get('total', 0)}"
    story.append(Paragraph(info_text, normal_style))
    story.append(Spacer(1, 12))

    # Таблица
    data = [['ID', 'Время', 'Пользователь', 'Действие', 'Описание']]

    for log in logs_data.get('items', [])[:100]:  # Ограничиваем 100 записями для PDF
        row = [
            str(log.get('id', '')),
            log.get('created_at', '')[:16] if log.get('created_at') else '',  # Обрезаем до даты и времени
            log.get('username', ''),
            log.get('action_type', ''),
            log.get('description', '')[:50] + '...' if len(log.get('description', '')) > 50 else log.get('description', ''),
        ]
        data.append(row)

    table = Table(data, colWidths=[0.5*inch, 1.2*inch, 1.2*inch, 0.8*inch, 3.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))

    story.append(table)

    # Сборка PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


def create_excel_response(buffer: BytesIO, filename: Optional[str] = None) -> Response:
    """Создает Flask Response для Excel файла."""
    if filename is None:
        filename = f'audit_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    return Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


def create_pdf_response(buffer: BytesIO, filename: Optional[str] = None) -> Response:
    """Создает Flask Response для PDF файла."""
    if filename is None:
        filename = f'audit_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'

    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

