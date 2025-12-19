# app/services/multi_position_processor.py
from copy import copy
import math
from typing import List, Dict, Any
import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from io import BytesIO

# Константы из excel_row_inserter.py
DATA_START_ROW = 10
INSERT_ROW_INDEX = 11
MAX_ROWS = 500


class MultiPositionProcessor:
    """Обработчик множественных позиций для Excel файлов на основе excel_row_inserter.py"""
    
    def __init__(self, template_path: str):
        self.template_path = template_path
    
    def _extract_days_from_payment_terms(self, payment_terms: str) -> int | None:
        """
        Извлекает число дней из текста условий оплаты.
        
        Ищет последнее упоминание числа с "дней", "день", "дня" в тексте,
        так как обычно срок указывается в конце условий оплаты.
        
        Args:
            payment_terms: Текст условий оплаты
            
        Returns:
            Число дней или None, если не найдено
        """
        if not payment_terms:
            return None
        
        import re
        
        # Паттерны для поиска числа перед словами "дней", "день", "дня"
        # Ищем все совпадения, чтобы взять последнее (обычно срок в конце)
        patterns = [
            r'(\d+)\s*(?:дней|день|дня|дн\.?)',  # "30 дней", "30 день", "30 дня", "30 дн."
            r'в\s*течении\s*(\d+)\s*(?:дней|день|дня|дн\.?)',  # "в течении 30 дней"
            r'в\s*течение\s*(\d+)\s*(?:дней|день|дня|дн\.?)',  # "в течение 30 дней"
            r'отсрочка\s*(?:в\s*течении|в\s*течение)?\s*(\d+)\s*(?:дней|день|дня|дн\.?)',  # "отсрочка в течении 10 дней"
            r'через\s*(\d+)\s*(?:дней|день|дня|дн\.?)',  # "через 30 дней"
        ]
        
        # Ищем все совпадения по каждому паттерну
        all_matches = []
        for pattern in patterns:
            matches = re.finditer(pattern, payment_terms, re.IGNORECASE)
            for match in matches:
                try:
                    days = int(match.group(1))
                    # Сохраняем позицию в тексте и значение
                    all_matches.append((match.start(), days))
                except (ValueError, IndexError):
                    continue
        
        # Если нашли совпадения, берем последнее (самое правое в тексте)
        if all_matches:
            # Сортируем по позиции и берем последнее
            all_matches.sort(key=lambda x: x[0])
            return all_matches[-1][1]
        
        # Если не нашли по паттернам, ищем просто числа в тексте
        # Но берем последнее число, так как срок обычно указывается в конце
        numbers = re.finditer(r'\d+', payment_terms)
        number_list = []
        for match in numbers:
            try:
                number_list.append((match.start(), int(match.group())))
            except (ValueError, IndexError):
                continue
        
        if number_list:
            # Берем последнее найденное число
            number_list.sort(key=lambda x: x[0])
            return number_list[-1][1]
        
        return None
        
    def shift_merged_cells(self, sheet, insert_row: int, rows_to_add: int) -> None:
        """Сдвигает объединенные ячейки вниз при добавлении строк."""
        merged_ranges = list(sheet.merged_cells.ranges)
        
        # Разъединяем все ячейки
        for merged_range in merged_ranges:
            sheet.unmerge_cells(str(merged_range))
        
        # Объединяем заново со сдвигом
        for merged_range in merged_ranges:
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            min_col = merged_range.min_col
            max_col = merged_range.max_col
            
            # Сдвигаем только те объединенные ячейки, которые находятся ниже места вставки
            if min_row >= insert_row:
                min_row += rows_to_add
                max_row += rows_to_add
            
            new_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            sheet.merge_cells(new_range)

    def shift_drawing_objects(self, sheet, insert_row: int, rows_to_add: int) -> None:
        """Сдвигает графические объекты (изображения, фигуры) вниз при добавлении строк."""
        if not hasattr(sheet, '_images'):
            return
            
        for image in sheet._images:
            anchor = image.anchor
            if hasattr(anchor, 'top') and anchor.top >= insert_row:
                anchor.top += rows_to_add
            if hasattr(anchor, 'bottom') and anchor.bottom >= insert_row:
                anchor.bottom += rows_to_add

    def ensure_sheet_capacity(self, sheet):
        """Гарантирует, что в листе достаточно строк для работы."""
        if sheet.max_row < MAX_ROWS:
            rows_to_add = MAX_ROWS - sheet.max_row
            sheet.insert_rows(sheet.max_row + 1, rows_to_add)

    def copy_row_style(self, sheet, source_row: int, target_row: int) -> None:
        """Копирует стили, формат чисел и высоту строки."""
        max_col = sheet.max_column
        for col in range(1, max_col + 1):
            src = sheet.cell(row=source_row, column=col)
            dst = sheet.cell(row=target_row, column=col)

            dst.value = None

            if src.has_style:
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.number_format = src.number_format
                dst.protection = copy(src.protection)
                dst.alignment = copy(src.alignment)

        if source_row in sheet.row_dimensions:
            source_dim = sheet.row_dimensions[source_row]
            target_dim = sheet.row_dimensions[target_row]
            target_dim.height = source_dim.height

    def copy_row_values(self, sheet, source_row: int, target_row: int) -> None:
        """Копирует значения и формулы, сдвигая относительные ссылки."""
        max_col = sheet.max_column

        for col in range(1, max_col + 1):
            src = sheet.cell(row=source_row, column=col)
            dst = sheet.cell(row=target_row, column=col)

            if src.value is None:
                dst.value = None
                continue

            if isinstance(src.value, str) and src.value.startswith("="):
                dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
            else:
                dst.value = src.value

    def find_last_data_row(self, sheet) -> int:
        """Возвращает последнюю строку, в которой в столбце B указано число."""
        for row in range(min(sheet.max_row, MAX_ROWS), DATA_START_ROW - 1, -1):
            cell_value = sheet[f"B{row}"].value
            if cell_value is None:
                continue
            try:
                int(cell_value)
                return row
            except (ValueError, TypeError):
                continue

        return DATA_START_ROW - 1

    def update_total_row_formulas(self, sheet) -> None:
        """Пересчитывает формулы в итоговой строке и строке с НДС."""
        last_data_row = max(self.find_last_data_row(sheet), DATA_START_ROW)
        if last_data_row < DATA_START_ROW:
            return

        total_row = last_data_row + 1

        total_formulas = {
            "I": f"=SUM(I{DATA_START_ROW}:I{last_data_row})",
            "O": f"=SUM(O{DATA_START_ROW}:O{last_data_row})",
            "Q": f"=SUM(Q{DATA_START_ROW}:Q{last_data_row})",
            "S": f"=SUM(S{DATA_START_ROW}:S{last_data_row})",
            "U": f"=SUM(U{DATA_START_ROW}:U{last_data_row})",
            "Y": f"=SUM(Y{DATA_START_ROW}:Y{last_data_row})",
        }

        for col, formula in total_formulas.items():
            sheet[f"{col}{total_row}"] = formula

        sheet[f"X{total_row}"] = f"=AVERAGE(X{DATA_START_ROW}:X{last_data_row})"
        # Итог в X также показываем как проценты без десятых
        sheet[f"X{total_row}"].number_format = '0%'
        sheet[f"Z{total_row}"] = (
            f"=(I{total_row}-O{total_row}-S{total_row}-U{total_row}-"
            f"Y{total_row}-AA{total_row}-AC{total_row}-AB{total_row})/I{total_row}"
        )

        nds_row = total_row + 1
        sheet[f"I{nds_row}"] = f"=I{total_row}*1.2"

    def update_summary_block(self, sheet, rows_added: int = 0) -> None:
        """Обновляет формулы итоговых строк c учётом новой строки."""
        last_data_row = max(self.find_last_data_row(sheet), DATA_START_ROW)
        total_row = last_data_row + 1
        
        # Динамически вычисляем все позиции на основе total_row
        k_formula_row = total_row + 4
        i20_row = total_row + 8
        i25_row = total_row + 13
        i26_row = i25_row + 1
        i27_row = i26_row + 1
        i28_row = i27_row + 1
        i29_row = i28_row + 1
        i30_row = i29_row + 1
        i31_row = i30_row + 1
        i32_row = i31_row + 1
        i33_row = i32_row + 1
        i34_row = i33_row + 1
        i35_row = i34_row + 1
        i36_row = i35_row + 1
        i37_row = i36_row + 1

        # Обновляем формулы с динамическими ссылками
        sheet[f"I{total_row}"] = f"=SUM(I{DATA_START_ROW}:I{last_data_row})"
        sheet[f"I{total_row + 1}"] = f"=I{total_row}*1.2"

        sheet[f"K{k_formula_row}"] = f"=I{k_formula_row}+I{k_formula_row + 1}"

        o_base_row = total_row
        sheet[f"O{o_base_row + 5}"] = f"=I{o_base_row}*14"
        sheet[f"O{o_base_row + 6}"] = f"=I{o_base_row}*14"

        # Формула в O17 должна динамически меняться
        o17_row = total_row + 6
        sheet[f"O{o17_row}"] = f"=I{total_row + 1}*14"

        sheet[f"I{i20_row}"] = f"=I{total_row}"

        sheet[f"I{i25_row}"] = f"=I{total_row + 1}"

        sheet[f"I{i26_row}"] = f"=I{i25_row}/120*20"
        sheet[f"I{i27_row}"] = f"=I{i25_row}-I{i26_row}"
        sheet[f"I{i28_row}"] = f"=SUM(I{i29_row}:I{i28_row + 10})"
        sheet[f"I{i29_row}"] = f"=O{total_row}"

        # Динамическая формула ЕСЛИ для строки I29/I30/... в зависимости от числа позиций
        positions_count = max(0, last_data_row - DATA_START_ROW + 1)
        # Формула должна всегда стоять в строке шаблона I29, которая динамически равна i30_row
        target_if_row = i30_row
        sheet[f"I{target_if_row}"] = (
            f"=IF(H{target_if_row}=D{target_if_row + 14},I{target_if_row - 1}*3.2%,0)"
        )
        
        # Формулы для I30-I33 ссылаются на итоги по соответствующим столбцам
        sheet[f"I{i31_row}"] = f"=Y{total_row}"
        sheet[f"I{i32_row}"] = f"=S{total_row}"
        sheet[f"I{i33_row}"] = f"=U{total_row}"

        sheet[f"I{i34_row}"] = f"=IF(H{i34_row}=\"ДА\",I{i29_row}*16%/365*K{k_formula_row},0)"
        sheet[f"I{i35_row}"] = f"=AA{total_row}"
        sheet[f"I{i36_row}"] = f"=AB{total_row}"
        sheet[f"I{i37_row}"] = f"=AC{total_row}"

        # Формула банковской гарантии
        bank_guarantee_row = 37 + rows_added
        sheet[f"I{bank_guarantee_row}"] = (
            f"=IF(H{bank_guarantee_row}=\"ДА\",I{i25_row}*3%/365*(I{k_formula_row}+I{k_formula_row + 1}),0)"
        )

        # Обновляем остальные формулы после банковской гарантии
        difference_row = bank_guarantee_row + 1
        ratio_row = difference_row + 2
        
        sheet[f"I{difference_row}"] = f"=I{i27_row}-I{i28_row}"
        sheet[f"I{ratio_row}"] = f"=I{difference_row}/I{i27_row}"

    def update_logistics_columns(self, sheet) -> None:
        """Настраивает формулы в столбцах R и T для всех строк с данными."""
        last_data_row = max(self.find_last_data_row(sheet), DATA_START_ROW)
        if last_data_row < DATA_START_ROW:
            return

        total_row = last_data_row + 1
        logistics_row = total_row + 3

        for row in range(DATA_START_ROW, last_data_row + 1):
            sheet[f"R{row}"] = f"=$U${logistics_row}/$Q${total_row}*P{row}/11.5*0.3"
            sheet[f"T{row}"] = f"=$U${logistics_row}/$Q${total_row}*P{row}/11.5*0.7"

    def update_row_numbers(self, sheet) -> None:
        """Обновляет нумерацию строк в столбце B начиная с DATA_START_ROW."""
        last_data_row = self.find_last_data_row(sheet)
        if last_data_row < DATA_START_ROW:
            return
        
        # Нумеруем все строки с данными начиная с 1
        for i, row in enumerate(range(DATA_START_ROW, last_data_row + 1), 1):
            sheet[f"B{row}"] = i

    def insert_new_rows(self, sheet, rows_to_add: int) -> None:
        """Вставляет новые строки в лист Excel."""
        # Гарантируем достаточную емкость листа
        self.ensure_sheet_capacity(sheet)

        # Сдвигаем объединенные ячейки и графические объекты
        self.shift_merged_cells(sheet, INSERT_ROW_INDEX, rows_to_add)
        self.shift_drawing_objects(sheet, INSERT_ROW_INDEX, rows_to_add)

        # Вставляем строки
        sheet.insert_rows(INSERT_ROW_INDEX, rows_to_add)

        # Копируем стили и значения для каждой новой строки
        for offset in range(rows_to_add):
            target_row = INSERT_ROW_INDEX + offset
            self.copy_row_style(sheet, DATA_START_ROW, target_row)
            self.copy_row_values(sheet, DATA_START_ROW, target_row)

        # Обновляем все формулы и нумерацию
        self.update_row_numbers(sheet)
        self.update_total_row_formulas(sheet)
        self.update_summary_block(sheet, rows_to_add)
        self.update_logistics_columns(sheet)

    def fill_position_data(self, sheet, position_data: Dict[str, Any], row_number: int) -> None:
        """Заполняет данные позиции в указанной строке."""
        # Маппинг полей формы на ячейки Excel
        field_mapping = {
            'product': 'C',  # Наименование товара (будет заполнено отдельно с номером чертежа)
            'drawing_number': 'E',  # Номер чертежа
            'material': 'D',  # Материал
            'cost_price': 'M',  # Сумма закупа
            'cost_price_per_kg': 'N',  # Цена за кг
            'quantity': 'G',  # Количество
            'weight': 'P',  # Вес за шт
            'duty_percent': 'X',  # Пошлина
        }
        
        # Сначала заполняем столбец C с объединенным форматом "продукт ч.номер чертежа"
        product = str(position_data.get('product', '')).strip()
        drawing_number = str(position_data.get('drawing_number', '')).strip()
        
        if product:
            if drawing_number:
                # Формируем строку в формате "продукт ч.номер чертежа"
                product_value = f"{product} ч.{drawing_number}"
            else:
                product_value = product
            sheet[f"C{row_number}"] = product_value
        
        # Заполняем остальные поля
        for field, column in field_mapping.items():
            # Пропускаем product, так как он уже обработан выше
            if field == 'product':
                continue
                
            if field in position_data and position_data[field]:
                value = position_data[field]
                if field in ['cost_price', 'cost_price_per_kg', 'weight', 'duty_percent']:
                    try:
                        value = float(value)
                        if field == 'duty_percent':
                            value = value / 100  # Конвертируем проценты в десятичную дробь
                    except (ValueError, TypeError):
                        value = 0
                elif field == 'quantity':
                    try:
                        value = int(value)
                    except (ValueError, TypeError):
                        value = 1
                
                cell = sheet[f"{column}{row_number}"]
                cell.value = value
                # Форматируем числа: X-столбец как проценты без десятых, G-столбец как целое число, остальные до двух знаков
                if isinstance(value, (int, float)):
                    if column == 'X':
                        cell.number_format = '0%'
                    elif column == 'G':
                        cell.number_format = '0'  # Целое число для количества
                    else:
                        cell.number_format = '0.00'

    def fill_common_data(self, sheet, form_data: Dict[str, Any], final_price: float = None, general_price: float = None, manager_fio: str = None) -> None:
        """Заполняет общие данные (компания, логистика, дата и т.д.) в Excel."""
        from datetime import datetime
        
        current_date = datetime.now().strftime('%d.%m.%Yг.')
        company = form_data.get('company', '').strip()
        logistics = float(form_data.get('logistics', 0))
        delivery_time = int(form_data.get('delivery_time', 0))
        tender_number = form_data.get('tender_number', '').strip()
        delivery_address = form_data.get('delivery_address', '').strip()
        payment_terms = form_data.get('payment_terms', '').strip()
        proposal_validity = form_data.get('proposal_validity', '').strip()
        
        # Извлекаем число дней из условий оплаты
        payment_days = self._extract_days_from_payment_terms(payment_terms)
        
        # Заполняем общие поля
        sheet['D4'] = company  # Название компании
        sheet['U14'] = logistics  # Логистика
        sheet['I15'] = delivery_time  # Срок поставки
        sheet['D2'] = current_date  # Дата
        sheet['D5'] = tender_number  # Номер тендера
        sheet['P4'] = delivery_address  # Адрес доставки
        # Условия оплаты: полный текст с префиксом в B16, число дней в I16
        if payment_terms:
            sheet['B16'] = f"Условия оплаты: {payment_terms}"
        if payment_days is not None:
            sheet['I16'] = payment_days
        # Срок действия предложения НЕ вставляется в Excel
        
        # Заполняем ФИО менеджера в ячейку N22
        if manager_fio:
            sheet['N22'] = manager_fio
        
        # Заполняем цены, если они переданы
        if final_price is not None:
            # Округляем вверх до десятков
            rounded_fp = math.ceil(float(final_price) / 10.0) * 10.0
            sheet['H10'] = rounded_fp  # Цена за единицу
        if general_price is not None:
            gp = round(float(general_price), 2)
            sheet['I11'] = gp  # Общая цена

    def process_multiple_positions(self, positions: List[Dict[str, Any]], form_data: Dict[str, Any] = None, final_price: float = None, general_price: float = None, position_prices: List[Dict[str, Any]] | None = None, manager_fio: str = None) -> BytesIO:
        """Обрабатывает множественные позиции и возвращает Excel файл в памяти."""
        if not positions:
            raise ValueError("Список позиций не может быть пустым")
        
        # Загружаем шаблон
        workbook = openpyxl.load_workbook(self.template_path)
        sheet = workbook.active
        
        # Заполняем общие данные, если они переданы
        if form_data:
            self.fill_common_data(sheet, form_data, final_price, general_price, manager_fio)
        
        # Количество позиций для добавления (минус 1, так как первая строка уже есть)
        positions_to_add = len(positions) - 1
        
        if positions_to_add > 0:
            # Вставляем дополнительные строки
            self.insert_new_rows(sheet, positions_to_add)
        
        # Заполняем данные для всех позиций
        for i, position in enumerate(positions):
            row_number = DATA_START_ROW + i
            sheet[f"B{row_number}"] = i + 1
            self.fill_position_data(sheet, position, row_number)

            # Проставляем рассчитанные цены по позиции, если переданы
            if position_prices and i < len(position_prices):
                pp = position_prices[i]
                fp = pp.get('final_price')
                if fp is not None:
                    # Округляем вверх до десятков
                    fp_rounded = math.ceil(float(fp) / 10.0) * 10.0
                    sheet[f"H{row_number}"] = fp_rounded  # Цена за единицу по позиции
                # Вставляем формулу для выручки: цена за единицу * количество
                sheet[f"I{row_number}"] = f"=H{row_number}*G{row_number}"
        
        # Гарантируем обновление формул и сводного блока даже при одной позиции
        self.update_total_row_formulas(sheet)
        self.update_summary_block(sheet, positions_to_add)
        self.update_logistics_columns(sheet)
        
        # Сохраняем в BytesIO
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        workbook.close()
        
        return excel_file
