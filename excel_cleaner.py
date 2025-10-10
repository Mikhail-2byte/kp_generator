import openpyxl
from openpyxl import load_workbook
import re

def process_budget_template():
    """
    Интерактивный процессор для бюджетного шаблона
    """
    input_file = "Бюджет шаблон.xlsx"
    output_file = "Бюджет_обработанный.xlsx"
    
    # Загружаем workbook
    wb = load_workbook(input_file)
    ws = wb['расчет']
    
    print("Файл загружен успешно")
    print("\n=== ЗАПОЛНЕНИЕ ПОЗИЦИЙ ===")
    
    # Запрашиваем наименования позиций у пользователя
    fill_positions(ws)
    
    # Определяем заполненные строки в таблице (10-30)
    start_row = 10
    end_row = 30
    filled_rows = []
    
    for row in range(start_row, end_row + 1):
        cell_c = ws[f'C{row}']
        if cell_c.value is not None and str(cell_c.value).strip() != '':
            filled_rows.append(row)
    
    print(f"Заполненные строки: {filled_rows}")
    
    if not filled_rows:
        print("Нет заполненных строк!")
        return
    
    # Удаляем пустые строки из таблицы (с конца чтобы не сбить нумерацию)
    empty_rows = [row for row in range(start_row, end_row + 1) if row not in filled_rows]
    
    for row in sorted(empty_rows, reverse=True):
        ws.delete_rows(row)
        print(f"Удалена пустая строка {row}")
    
    # Теперь обновляем ВСЕ формулы в строке 31 чтобы они ссылались на новый диапазон
    new_start_row = min(filled_rows)
    new_end_row = max(filled_rows)
    
    print(f"\n=== ОБНОВЛЕНИЕ ФОРМУЛ В СТРОКЕ 31 ===")
    print(f"Новый диапазон данных: строки {new_start_row}-{new_end_row}")
    
    # Обновляем формулы в строке 31
    update_row_31_formulas(ws, new_start_row, new_end_row)
    
    # Обновляем формулы логистики в оставшихся строках
    update_logistics_formulas(ws, filled_rows)
    
    # Сохраняем результат
    wb.save(output_file)
    print(f"\nОбработанный файл сохранен как: {output_file}")
    print("Готово!")

def fill_positions(ws):
    """
    Запрашивает у пользователя наименования позиций и заполняет столбец C
    """
    print("Введите наименования позиций (для завершения введите 'готово'):")
    
    row = 10
    position_count = 1
    
    while row <= 30:
        position_name = input(f"Позиция {position_count}: ").strip()
        
        if position_name.lower() == 'готово':
            print("Завершение ввода...")
            break
            
        if position_name:  # Если введен не пустой текст
            # Заполняем ячейку в столбце C
            ws[f'C{row}'] = position_name
            print(f"  → Записано в строку {row}")
            row += 1
            position_count += 1
        else:
            print("  → Пропуск пустой позиции")
            break
    
    print(f"Всего заполнено позиций: {position_count - 1}")

def update_row_31_formulas(ws, start_row, end_row):
    """
    Обновляет ВСЕ формулы в строке 31 чтобы они ссылались на новый диапазон
    """
    # Основные формулы SUM в строке 31
    sum_formulas = {
        'I31': f'=SUM(I{start_row}:I{end_row})',
        'O31': f'=SUM(O{start_row}:O{end_row})', 
        'Q31': f'=SUM(Q{start_row}:Q{end_row})',
        'S31': f'=SUM(S{start_row}:S{end_row})',
        'U31': f'=SUM(U{start_row}:U{end_row})',
        'Y31': f'=SUM(Y{start_row}:Y{end_row})',
        'X31': f'=AVERAGE(X{start_row}:X{end_row})'
    }
    
    for cell_ref, new_formula in sum_formulas.items():
        if cell_ref in ws:
            ws[cell_ref] = new_formula
            print(f"Обновлена формула {cell_ref}: {new_formula}")
    
    # Формула маржи в Z31
    margin_formula = f'=(I31-O31-S31-U31-Y31-AA31-AC31-AB31)/I31'
    if 'Z31' in ws:
        ws['Z31'] = margin_formula
        print(f"Обновлена формула Z31: {margin_formula}")
    
    # Формула ИТОГО с НДС в I32
    if 'I32' in ws:
        ws['I32'] = '=I31*1.2'
        print("Обновлена формула I32: =I31*1.2")

def update_logistics_formulas(ws, filled_rows):
    """
    Обновляет формулы логистики в оставшихся строках
    """
    print(f"\n=== ОБНОВЛЕНИЕ ФОРМУЛ ЛОГИСТИКИ ===")
    
    for row in filled_rows:
        # Формулы логистики КНР
        r_formula = f'=$U$31/Q31*P{row}/12*0.3'
        t_formula = f'=$U$31/Q31*P{row}/12*0.7'
        
        ws[f'R{row}'] = r_formula
        ws[f'T{row}'] = t_formula
        
        print(f"Обновлены формулы логистики в строке {row}")

def update_all_references(ws, old_start, old_end, new_start, new_end):
    """
    Обновляет все ссылки в формулах на старый диапазон
    """
    print(f"\n=== ОБНОВЛЕНИЕ ВСЕХ ССЫЛОК ===")
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            
            if cell.data_type == 'f' and cell.value:  # Если это формула
                formula = cell.value
                
                # Заменяем старый диапазон на новый во всех формулах
                old_range = f"I{old_start}:I{old_end}"
                new_range = f"I{new_start}:I{new_end}"
                formula = formula.replace(old_range, new_range)
                
                old_range = f"O{old_start}:O{old_end}"
                new_range = f"O{new_start}:O{new_end}"
                formula = formula.replace(old_range, new_range)
                
                old_range = f"Q{old_start}:Q{old_end}"
                new_range = f"Q{new_start}:Q{new_end}"
                formula = formula.replace(old_range, new_range)
                
                old_range = f"S{old_start}:S{old_end}"
                new_range = f"S{new_start}:S{new_end}"
                formula = formula.replace(old_range, new_range)
                
                old_range = f"U{old_start}:U{old_end}"
                new_range = f"U{new_start}:U{new_end}"
                formula = formula.replace(old_range, new_range)
                
                old_range = f"Y{old_start}:Y{old_end}"
                new_range = f"Y{new_start}:Y{new_end}"
                formula = formula.replace(old_range, new_range)
                
                old_range = f"X{old_start}:X{old_end}"
                new_range = f"X{new_start}:X{new_end}"
                formula = formula.replace(old_range, new_range)
                
                # Если формула изменилась, обновляем ячейку
                if formula != cell.value:
                    cell.value = formula
                    print(f"Обновлена ссылка в {cell.coordinate}")

if __name__ == "__main__":
    try:
        process_budget_template()
        
        # Даем пользователю возможность просмотреть результат
        input("\nНажмите Enter для завершения...")
        
    except FileNotFoundError:
        print("Ошибка: Файл 'Бюджет шаблон.xlsx' не найден в текущей папке")
        input("Нажмите Enter для завершения...")
    except Exception as e:
        print(f"Произошла ошибка: {e}")
        import traceback
        traceback.print_exc()
        input("Нажмите Enter для завершения...")