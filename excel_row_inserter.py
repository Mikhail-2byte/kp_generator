from copy import copy
from pathlib import Path
import shutil

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter


TEMPLATE_FILE = Path("Бюджет шаблон.xlsx")
OUTPUT_FILE = Path("Обработанный_добавление.xlsx")
DATA_START_ROW = 10
INSERT_ROW_INDEX = 11
DEFAULT_ROWS_TO_ADD = 1
MAX_ROWS = 500


def shift_merged_cells(sheet, insert_row: int, rows_to_add: int) -> None:
    """Сдвигает объединенные ячейки вниз при добавлении строк."""
    merged_ranges = list(sheet.merged_cells.ranges)
    
    # Разъединяем все ячейки
    for merged_range in merged_ranges:
        sheet.unmerge_cells(str(merged_range))
    
    # Объединяем заново со сдвигом
    for merged_range in merged_ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col
        
        # Сдвигаем только те объединенные ячейки, которые находятся ниже места вставки
        if min_row >= insert_row:
            min_row += rows_to_add
            max_row += rows_to_add
        
        new_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        sheet.merge_cells(new_range)


def shift_drawing_objects(sheet, insert_row: int, rows_to_add: int) -> None:
    """Сдвигает графические объекты (изображения, фигуры) вниз при добавлении строк."""
    if not hasattr(sheet, '_images'):
        return
        
    for image in sheet._images:
        anchor = image.anchor
        if hasattr(anchor, 'top') and anchor.top >= insert_row:
            anchor.top += rows_to_add
        if hasattr(anchor, 'bottom') and anchor.bottom >= insert_row:
            anchor.bottom += rows_to_add


def ensure_sheet_capacity(sheet):
    """Гарантирует, что в листе достаточно строк для работы."""
    if sheet.max_row < MAX_ROWS:
        rows_to_add = MAX_ROWS - sheet.max_row
        sheet.insert_rows(sheet.max_row + 1, rows_to_add)


def copy_row_style(sheet, source_row: int, target_row: int) -> None:
    """Копирует стили, формат чисел и высоту строки."""

    max_col = sheet.max_column
    for col in range(1, max_col + 1):
        src = sheet.cell(row=source_row, column=col)
        dst = sheet.cell(row=target_row, column=col)

        dst.value = None

        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)

    if source_row in sheet.row_dimensions:
        source_dim = sheet.row_dimensions[source_row]
        target_dim = sheet.row_dimensions[target_row]
        target_dim.height = source_dim.height


def copy_row_values(sheet, source_row: int, target_row: int) -> None:
    """Копирует значения и формулы, сдвигая относительные ссылки."""

    max_col = sheet.max_column

    for col in range(1, max_col + 1):
        src = sheet.cell(row=source_row, column=col)
        dst = sheet.cell(row=target_row, column=col)

        if src.value is None:
            dst.value = None
            continue

        if isinstance(src.value, str) and src.value.startswith("="):
            dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
        else:
            dst.value = src.value


def find_last_data_row(sheet) -> int:
    """Возвращает последнюю строку, в которой в столбце B указано число."""

    for row in range(min(sheet.max_row, MAX_ROWS), DATA_START_ROW - 1, -1):
        cell_value = sheet[f"B{row}"].value
        if cell_value is None:
            continue
        try:
            int(cell_value)
            return row
        except (ValueError, TypeError):
            continue

    return DATA_START_ROW - 1


def update_total_row_formulas(sheet) -> None:
    """Пересчитывает формулы в итоговой строке и строке с НДС."""

    last_data_row = max(find_last_data_row(sheet), INSERT_ROW_INDEX)
    if last_data_row < DATA_START_ROW:
        return

    total_row = last_data_row + 1

    total_formulas = {
        "I": f"=SUM(I{DATA_START_ROW}:I{last_data_row})",
        "O": f"=SUM(O{DATA_START_ROW}:O{last_data_row})",
        "Q": f"=SUM(Q{DATA_START_ROW}:Q{last_data_row})",
        "S": f"=SUM(S{DATA_START_ROW}:S{last_data_row})",
        "U": f"=SUM(U{DATA_START_ROW}:U{last_data_row})",
        "Y": f"=SUM(Y{DATA_START_ROW}:Y{last_data_row})",
    }

    for col, formula in total_formulas.items():
        sheet[f"{col}{total_row}"] = formula

    sheet[f"X{total_row}"] = f"=AVERAGE(X{DATA_START_ROW}:X{last_data_row})"
    sheet[f"Z{total_row}"] = (
        f"=(I{total_row}-O{total_row}-S{total_row}-U{total_row}-"
        f"Y{total_row}-AA{total_row}-AC{total_row}-AB{total_row})/I{total_row}"
    )

    nds_row = total_row + 1
    sheet[f"I{nds_row}"] = f"=I{total_row}*1.2"


def update_summary_block(sheet, rows_added: int = 0) -> None:
    """Обновляет формулы итоговых строк c учётом новой строки."""

    last_data_row = max(find_last_data_row(sheet), INSERT_ROW_INDEX)
    total_row = last_data_row + 1
    
    # Динамически вычисляем все позиции на основе total_row
    k_formula_row = total_row + 4
    i20_row = total_row + 8
    i25_row = total_row + 13
    i26_row = i25_row + 1
    i27_row = i26_row + 1
    i28_row = i27_row + 1
    i29_row = i28_row + 1
    i30_row = i29_row + 1
    i31_row = i30_row + 1
    i32_row = i31_row + 1
    i33_row = i32_row + 1
    i34_row = i33_row + 1
    i35_row = i34_row + 1
    i36_row = i35_row + 1
    i37_row = i36_row + 1

    # Обновляем формулы с динамическими ссылками
    sheet[f"I{total_row}"] = f"=SUM(I{DATA_START_ROW}:I{last_data_row})"
    sheet[f"I{total_row + 1}"] = f"=I{total_row}*1.2"

    sheet[f"K{k_formula_row}"] = f"=I{k_formula_row}+I{k_formula_row + 1}"

    o_base_row = total_row
    sheet[f"O{o_base_row + 5}"] = f"=I{o_base_row}*14"
    sheet[f"O{o_base_row + 6}"] = f"=I{o_base_row}*14"

    # Формула в O17 должна динамически меняться
    o17_row = total_row + 6
    sheet[f"O{o17_row}"] = f"=I{total_row + 1}*14"

    sheet[f"I{i20_row}"] = f"=I{total_row}"

    sheet[f"I{i25_row}"] = f"=I{total_row + 1}"

    sheet[f"I{i26_row}"] = f"=I{i25_row}/120*20"
    sheet[f"I{i27_row}"] = f"=I{i25_row}-I{i26_row}"
    sheet[f"I{i28_row}"] = f"=SUM(I{i29_row}:I{i28_row + 10})"

    sheet[f"I{i29_row}"] = f"=O{total_row}"
    
    d_reference_row = total_row + 32
    sheet[f"I{i30_row}"] = f"=IF(H{i30_row}=D{d_reference_row},I{i29_row}*3.2%,0)"

    sheet[f"I{i31_row}"] = f"=Y{total_row}"
    sheet[f"I{i32_row}"] = f"=S{total_row}"
    sheet[f"I{i33_row}"] = f"=U{total_row}"
    sheet[f"I{i34_row}"] = f"=IF(H{i34_row}=\"ДА\",I{i29_row}*16%/365*K{k_formula_row},0)"
    sheet[f"I{i35_row}"] = f"=AA{total_row}"
    sheet[f"I{i36_row}"] = f"=AB{total_row}"
    sheet[f"I{i37_row}"] = f"=AC{total_row}"

    # Формула банковской гарантии
    bank_guarantee_row = 37 + rows_added
    sheet[f"I{bank_guarantee_row}"] = (
        f"=IF(H{bank_guarantee_row}=\"ДА\",I{i25_row}*3%/365*(I{k_formula_row}+I{k_formula_row + 1}),0)"
    )

    # Обновляем остальные формулы после банковской гарантии
    difference_row = bank_guarantee_row + 1
    ratio_row = difference_row + 2
    
    sheet[f"I{difference_row}"] = f"=I{i27_row}-I{i28_row}"
    sheet[f"I{ratio_row}"] = f"=I{difference_row}/I{i27_row}"


def update_logistics_columns(sheet) -> None:
    """Настраивает формулы в столбцах R и T для всех строк с данными."""

    last_data_row = max(find_last_data_row(sheet), INSERT_ROW_INDEX)
    if last_data_row < DATA_START_ROW:
        return

    total_row = last_data_row + 1
    logistics_row = total_row + 3

    for row in range(DATA_START_ROW, last_data_row + 1):
        sheet[f"R{row}"] = f"=$U${logistics_row}/$Q${total_row}*P{row}/12*0.3"
        sheet[f"T{row}"] = f"=$U${logistics_row}/$Q${total_row}*P{row}/12*0.7"


def update_row_numbers(sheet) -> None:
    """Обновляет нумерацию строк в столбце B начиная с DATA_START_ROW."""
    
    last_data_row = find_last_data_row(sheet)
    if last_data_row < DATA_START_ROW:
        return
    
    # Нумеруем все строки с данными начиная с 1
    for i, row in enumerate(range(DATA_START_ROW, last_data_row + 1), start=1):
        sheet[f"B{row}"] = i


def prompt_rows_to_add() -> int:
    while True:
        raw_value = input(
            f"Введите количество строк для добавления (по умолчанию {DEFAULT_ROWS_TO_ADD}, максимум {MAX_ROWS - DATA_START_ROW}): "
        ).strip()

        if not raw_value:
            return DEFAULT_ROWS_TO_ADD

        try:
            rows = int(raw_value)
            if rows < 1:
                print("Ошибка: введите целое число больше нуля.")
                continue
                
            max_allowed = MAX_ROWS - DATA_START_ROW
            if rows > max_allowed:
                print(f"Ошибка: невозможно добавить более {max_allowed} строк.")
                continue
                
            return rows
        except ValueError:
            print("Ошибка: введите целое число больше нуля.")


def insert_new_rows(rows_to_add: int) -> None:
    if not TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"Файл '{TEMPLATE_FILE}' не найден")

    shutil.copy2(TEMPLATE_FILE, OUTPUT_FILE)

    workbook = openpyxl.load_workbook(OUTPUT_FILE)
    sheet = workbook.active

    # Гарантируем достаточную емкость листа
    ensure_sheet_capacity(sheet)

    # ОБНОВЛЕНИЕ: Сдвигаем объединенные ячейки и графические объекты
    shift_merged_cells(sheet, INSERT_ROW_INDEX, rows_to_add)
    shift_drawing_objects(sheet, INSERT_ROW_INDEX, rows_to_add)

    # Вставляем строки
    sheet.insert_rows(INSERT_ROW_INDEX, rows_to_add)

    # Копируем стили и значения для каждой новой строки
    for offset in range(rows_to_add):
        target_row = INSERT_ROW_INDEX + offset
        copy_row_style(sheet, DATA_START_ROW, target_row)
        copy_row_values(sheet, DATA_START_ROW, target_row)

    # Обновляем все формулы и нумерацию
    update_row_numbers(sheet)
    update_total_row_formulas(sheet)
    update_summary_block(sheet, rows_to_add)
    update_logistics_columns(sheet)

    # Оптимизируем производительность для больших файлов
    if rows_to_add > 50:
        workbook.save(OUTPUT_FILE)
        print(f"Обработано {rows_to_add} строк. Сохраняем файл...")
    
    workbook.save(OUTPUT_FILE)
    workbook.close()

    print(
        f"Добавлено строк: {rows_to_add}. Форматы перенесены, формулы обновлены. "
        f"Результат сохранён в '{OUTPUT_FILE}'."
    )


if __name__ == "__main__":
    rows_to_add = prompt_rows_to_add()
    insert_new_rows(rows_to_add)