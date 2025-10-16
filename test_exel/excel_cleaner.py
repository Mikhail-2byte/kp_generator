import openpyxl
import shutil
from pathlib import Path
import re

def process_budget_template():
    # Определяем пути к файлам
    source_file = Path("Бюджет шаблон.xlsx")
    destination_file = Path("Обработанный.xlsx")
    
    try:
        # Копируем файл с помощью shutil
        shutil.copy2(source_file, destination_file)
        print(f"Файл успешно обработан и сохранен как '{destination_file}'")
        
        # Открываем файл для редактирования
        workbook = openpyxl.load_workbook(destination_file)
        sheet = workbook.active
        
        # Сохраняем исходную высоту строки 31 для последующего использования
        original_row_31_height = sheet.row_dimensions[31].height if 31 in sheet.row_dimensions else None
        
        # Показываем текущие данные
        print(f"\nТекущие данные в файле:")
        print(f"Всего строк: {sheet.max_row}")
        
        # Запрашиваем у пользователя номер строки для удаления
        while True:
            try:
                row_to_delete = int(input("\nВведите номер строки для удаления (или 0 для выхода): "))
                
                if row_to_delete == 0:
                    print("Выход из программы.")
                    break
                elif 1 <= row_to_delete <= sheet.max_row:
                    # Сохраняем формулы из удаляемой строки для отладки
                    if row_to_delete <= sheet.max_row:
                        print(f"Формулы в удаляемой строке {row_to_delete}:")
                        for col in ['I', 'J', 'L', 'N', 'O', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'Y', 'Z']:
                            cell = f"{col}{row_to_delete}"
                            formula = sheet[cell].value
                            print(f"  {cell}: {formula}")
                    
                    # Удаляем строку
                    sheet.delete_rows(row_to_delete)
                    
                    # Обновляем формулы во всех строках после удаления
                    update_all_formulas(sheet, row_to_delete, original_row_31_height)
                    
                    workbook.save(destination_file)
                    print(f"Строка {row_to_delete} успешно удалена и все формулы обновлены!")
                    print(f"Теперь всего строк: {sheet.max_row}")
                    
                    # Предлагаем удалить еще одну строку
                    continue_deleting = input("Хотите удалить еще одну строку? (y/n): ").lower()
                    if continue_deleting != 'y':
                        break
                else:
                    print(f"Ошибка: Номер строки должен быть от 1 до {sheet.max_row}")
                    
            except ValueError:
                print("Ошибка: Введите корректный номер строки (целое число)")
            except Exception as e:
                print(f"Произошла ошибка при удаления строки: {str(e)}")
        
        workbook.close()
        print(f"\nФайл '{destination_file}' успешно обновлен!")
        
    except FileNotFoundError:
        print(f"Ошибка: Файл '{source_file}' не найден")
    except Exception as e:
        print(f"Произошла ошибка при обработке файла: {str(e)}")

def update_all_formulas(sheet, deleted_row, original_row_31_height=None):
    """
    Обновляет формулы во всех строках после удаления строки
    """
    # Определяем последнюю строку с данными в столбце B (номера п/п)
    last_data_row = find_last_data_row(sheet)
    print(f"Последняя строка с данными: {last_data_row}")
    
    # Определяем новую позицию итоговой строки (бывшая 31)
    total_row_position = last_data_row + 1
    
    # Обновляем формулы в столбцах I, J, L, N, O, Q, R, T, V, W, Y, Z для строк с данными
    for row in range(10, last_data_row + 1):
        for col in ['I', 'J', 'L', 'N', 'O', 'Q', 'R', 'T', 'V', 'W', 'Y', 'Z']:
            cell = f"{col}{row}"
            if sheet[cell].value and str(sheet[cell].value).startswith('='):
                formula = sheet[cell].value
                updated_formula = update_formula_references(formula, deleted_row)
                sheet[cell] = updated_formula
                print(f"Обновлена формула в {cell}: {updated_formula}")
    
    # Обновляем формулы в столбцах S и U только для строк с данными
    for row in range(10, last_data_row + 1):
        # Обновляем формулу в столбце S: =R<row>*L<row>
        s_cell = f"S{row}"
        sheet[s_cell] = f"=R{row}*L{row}"
        print(f"Установлена формула в {s_cell}: =R{row}*L{row}")
        
        # Обновляем формулу в столбце U: =T<row>*L<row>
        u_cell = f"U{row}"
        sheet[u_cell] = f"=T{row}*L{row}"
        print(f"Установлена формула в {u_cell}: =T{row}*L{row}")
    
    # Обновляем формулы в итоговой строке (теперь на позиции total_row_position)
    update_total_row_formulas(sheet, last_data_row, total_row_position)
    
    # Обновляем ячейку I39 (теперь она должна быть на позиции total_row_position + 8)
    update_i39_formula(sheet, total_row_position)
    
    # Обновляем ячейки во второй таблице
    update_second_table_formulas(sheet, total_row_position)
    
    # Обновляем ячейку K35 (теперь она должна быть на позиции total_row_position + 4)
    update_k35_formula(sheet, total_row_position)
    
    # Обновляем ячейки O36 и O37 (теперь они должны быть на позициях total_row_position + 5 и total_row_position + 6)
    update_o36_o37_formulas(sheet, total_row_position)
    
    # Настраиваем высоту строк после переноса
    adjust_row_heights(sheet, last_data_row, total_row_position, original_row_31_height)

def update_i39_formula(sheet, total_row_position):
    """
    Обновляет формулу в ячейке I39, которая теперь должна быть на 8 строк ниже итоговой
    """
    try:
        # Новая позиция ячейки I39 - на 8 строк ниже итоговой строки
        new_i39_row = total_row_position + 8
        
        # Устанавливаем новую формулу, которая ссылается на итоговую строку
        new_formula = f"=I{total_row_position}"
        
        # Устанавливаем формулу в ячейку I39
        sheet[f'I{new_i39_row}'] = new_formula
        print(f"Обновлена формула в I{new_i39_row}: {new_formula}")
        
        # Обновляем подпись если нужно
        if sheet[f'H{new_i39_row}'].value and "Итого по спецификации" in str(sheet[f'H{new_i39_row}'].value):
            sheet[f'H{new_i39_row}'] = "Итого по спецификации"
            
    except Exception as e:
        print(f"Ошибка при обновлении ячейки I39: {str(e)}")

def update_second_table_formulas(sheet, total_row_position):
    """
    Обновляет формулы во второй таблице (Итого по бюджету сделки)
    """
    try:
        # Находим начало второй таблицы по заголовку
        second_table_start = None
        for row in range(total_row_position + 10, min(total_row_position + 20, sheet.max_row + 1)):
            if sheet[f'B{row}'].value and "Итого по бюджету сделки" in str(sheet[f'B{row}'].value):
                second_table_start = row
                break
        
        if not second_table_start:
            print("Вторая таблица не найдена")
            return
            
        print(f"Начало второй таблицы: строка {second_table_start}")
        
        # Находим строки с "Прогноз выручки по контракту" и "Сумма НДС"
        revenue_row = None
        vat_row = None
        
        for row in range(second_table_start + 1, min(second_table_start + 10, sheet.max_row + 1)):
            cell_value = sheet[f'B{row}'].value
            if cell_value:
                if "Прогноз выручки по контракту" in str(cell_value):
                    revenue_row = row
                elif "Сумма НДС" in str(cell_value):
                    vat_row = row
        
        # Обновляем формулы
        if revenue_row:
            sheet[f'I{revenue_row}'] = f"=I{total_row_position}"
            print(f"Обновлена формула в I{revenue_row} (Прогноз выручки): =I{total_row_position}")
        
        if vat_row:
            sheet[f'I{vat_row}'] = f"=I{revenue_row}/120*20" if revenue_row else f"=I{total_row_position}/120*20"
            print(f"Обновлена формула в I{vat_row} (Сумма НДС): {sheet[f'I{vat_row}'].value}")
            
    except Exception as e:
        print(f"Ошибка при обновлении второй таблицы: {str(e)}")

def update_k35_formula(sheet, total_row_position):
    """
    Обновляет формулу в ячейке K35, которая теперь должна быть на 4 строки ниже итоговой
    """
    try:
        # Новая позиция ячейки K35 - на 4 строки ниже итоговой строки
        new_k35_row = total_row_position + 4
        
        # Находим строки с условиями поставки и оплаты
        delivery_row = None
        payment_row = None
        
        # Ищем строки с соответствующими подписями
        for row in range(total_row_position + 1, sheet.max_row + 1):
            cell_value = sheet[f'B{row}'].value
            if cell_value and "Условия Поставки" in str(cell_value):
                delivery_row = row
            elif cell_value and "Условия оплаты" in str(cell_value):
                payment_row = row
        
        # Если нашли обе строки, обновляем формулу
        if delivery_row and payment_row:
            # Формула должна ссылаться на ячейки I в найденных строках
            new_formula = f"=I{delivery_row}+I{payment_row}"
            sheet[f'K{new_k35_row}'] = new_formula
            print(f"Обновлена формула в K{new_k35_row}: {new_formula}")
        else:
            # Если не нашли, ищем по относительному положению
            delivery_row = total_row_position + 1
            payment_row = total_row_position + 3
            if delivery_row <= sheet.max_row and payment_row <= sheet.max_row:
                new_formula = f"=I{delivery_row}+I{payment_row}"
                sheet[f'K{new_k35_row}'] = new_formula
                print(f"Обновлена формула в K{new_k35_row} (относительное): {new_formula}")
            
    except Exception as e:
        print(f"Ошибка при обновлении ячейки K35: {str(e)}")

def update_o36_o37_formulas(sheet, total_row_position):
    """
    Обновляет формулы в ячейках O36 и O37, которые теперь должны быть на 5 и 6 строк ниже итоговой
    """
    try:
        # Новые позиции ячеек O36 и O37
        new_o36_row = total_row_position + 5
        new_o37_row = total_row_position + 6
        
        # Обновляем формулу в O36: должна ссылаться на итоговую строку без НДС
        o36_formula = f"=I{total_row_position}*14"
        sheet[f'O{new_o36_row}'] = o36_formula
        print(f"Обновлена формула в O{new_o36_row}: {o36_formula}")
        
        # Обновляем формулу в O37: должна ссылаться на итоговую строку с НДС
        o37_formula = f"=I{total_row_position+1}*14"
        sheet[f'O{new_o37_row}'] = o37_formula
        print(f"Обновлена формула в O{new_o37_row}: {o37_formula}")
        
        # Обновляем подписи если нужно
        if sheet[f'B{new_o36_row}'].value and "Итого без ндс" in str(sheet[f'B{new_o36_row}'].value):
            sheet[f'B{new_o36_row}'] = "Итого без ндс"
            
        if sheet[f'B{new_o37_row}'].value and "Итого с ндс" in str(sheet[f'B{new_o37_row}'].value):
            sheet[f'B{new_o37_row}'] = "Итого с ндс"
            
    except Exception as e:
        print(f"Ошибка при обновлении ячеек O36 и O37: {str(e)}")

def adjust_row_heights(sheet, last_data_row, total_row_position, original_row_31_height):
    """
    Настраивает высоту строк после удаления и переноса
    """
    # Уменьшаем высоту строки, которая была на позиции 31 (если она еще существует)
    if last_data_row + 2 <= sheet.max_row:
        # Строка, которая была на позиции 31, теперь на позиции last_data_row + 2
        old_row_31_position = last_data_row + 2
        if old_row_31_position in sheet.row_dimensions:
            sheet.row_dimensions[old_row_31_position].height = 15  # Стандартная высота
            print(f"Уменьшена высота строки {old_row_31_position}")
    
    # Устанавливаем нормальную высоту для новой итоговой строки
    if total_row_position in sheet.row_dimensions:
        sheet.row_dimensions[total_row_position].height = original_row_31_height if original_row_31_height else 20
        print(f"Установлена высота строки {total_row_position}: {sheet.row_dimensions[total_row_position].height}")
    
    # Устанавливаем нормальную высоту для строки с НДС
    nds_row = total_row_position + 1
    if nds_row in sheet.row_dimensions:
        sheet.row_dimensions[nds_row].height = 15
        print(f"Установлена высота строки {nds_row}: 15")

def update_total_row_formulas(sheet, last_data_row, total_row_position):
    """
    Обновляет формулы в итоговой строке (теперь на новой позиции)
    """
    if last_data_row < 10:
        return  # Нет данных для суммирования
    
    # Обновляем формулы суммирования
    total_formulas = {
        'I': f'=SUM(I10:I{last_data_row})',  # Сумма без НДС
        'O': f'=SUM(O10:O{last_data_row})',  # Сумма, без НДС (поставщик)
        'Q': f'=SUM(Q10:Q{last_data_row})',  # Итого вес, кг
        'S': f'=SUM(S10:S{last_data_row})',  # ИТОГО Логистики КНР
        'U': f'=SUM(U10:U{last_data_row})',  # ИТОГО Логистика РФ
        'Y': f'=SUM(Y10:Y{last_data_row})',  # Пошлина (сумма),
    }
    
    for col, formula in total_formulas.items():
        cell = f"{col}{total_row_position}"
        sheet[cell] = formula
        print(f"Обновлена формула в {cell}: {formula}")
    
    # Обновляем формулу среднего значения
    sheet[f'X{total_row_position}'] = f'=AVERAGE(X10:X{last_data_row})'
    print(f"Обновлена формула в X{total_row_position}: =AVERAGE(X10:X{last_data_row})")
    
    # Обновляем сложную формулу в Z (Маржа)
    z_formula = f'=(I{total_row_position}-O{total_row_position}-S{total_row_position}-U{total_row_position}-Y{total_row_position}-AA{total_row_position}-AC{total_row_position}-AB{total_row_position})/I{total_row_position}'
    sheet[f'Z{total_row_position}'] = z_formula
    print(f"Обновлена формула в Z{total_row_position}: {z_formula}")
    
    # Обновляем формулу ИТОГО с НДС (теперь на total_row_position + 1)
    nds_row = total_row_position + 1
    sheet[f'I{nds_row}'] = f'=I{total_row_position}*1.2'
    print(f"Обновлена формула в I{nds_row}: =I{total_row_position}*1.2")
    
    # Добавляем подпись "ИТОГО с НДС" если она была в исходном файле
    if sheet['H32'].value and "ИТОГО с НДС" in str(sheet['H32'].value):
        sheet[f'H{nds_row}'] = "ИТОГО с НДС"

def find_last_data_row(sheet):
    """
    Находит последнюю строку с данными в таблице
    Ищем строки с номерами п/п в столбце B
    """
    last_row = sheet.max_row
    # Ищем снизу вверх первую строку с номером п/п в столбце B
    for row in range(last_row, 0, -1):
        if sheet[f"B{row}"].value is not None:
            # Проверяем, что это число (номер п/п)
            try:
                int(sheet[f"B{row}"].value)
                return row
            except (ValueError, TypeError):
                continue
    return 29  # Возвращаем значение по умолчанию, если не нашли

def update_formula_references(formula, deleted_row):
    """
    Обновляет ссылки в формуле после удаления строки
    ВСЕ ссылки на строки >= deleted_row уменьшаются на 1
    """
    # Регулярное выражение для поиска ссылок на ячейки
    cell_ref_pattern = r'(\$?[A-Z]+\$?)(\d+)'
    
    def update_cell_ref(match):
        col_letter = match.group(1)  # Буква столбца (может быть с $)
        row_num = int(match.group(2))  # Номер строки
        
        # ВСЕ ссылки на строки >= deleted_row уменьшаем на 1
        if row_num >= deleted_row:
            return f"{col_letter}{row_num - 1}"
        else:
            # Ссылка на строку выше удаленной - оставляем как есть
            return f"{col_letter}{row_num}"
    
    # Обновляем все ссылки в формуле
    updated_formula = re.sub(cell_ref_pattern, update_cell_ref, formula)
    return updated_formula

def show_detailed_preview():
    """Функция для детального просмотра формул"""
    try:
        workbook = openpyxl.load_workbook("Обработанный.xlsx")
        sheet = workbook.active
        
        # Определяем последнюю строку с данными
        last_data_row = find_last_data_row(sheet)
        total_row_position = last_data_row + 1
        
        print(f"\nДетальный просмотр формул:")
        print(f"Последняя строка с данными: {last_data_row}")
        print(f"Итоговая строка: {total_row_position}")
        
        # Находим текущую позицию ячейки I39
        i39_position = total_row_position + 8
        if i39_position <= sheet.max_row:
            print(f"Текущая позиция ячейки I39: строка {i39_position}")
            print(f"Формула в I{i39_position}: {sheet[f'I{i39_position}'].value}")
        
        # Находим вторую таблицу
        second_table_start = None
        revenue_row = None
        vat_row = None
        
        for row in range(total_row_position + 10, min(total_row_position + 20, sheet.max_row + 1)):
            cell_value = sheet[f'B{row}'].value
            if cell_value:
                if "Итого по бюджету сделки" in str(cell_value):
                    second_table_start = row
                elif "Прогноз выручки по контракту" in str(cell_value):
                    revenue_row = row
                elif "Сумма НДС" in str(cell_value):
                    vat_row = row
        
        if second_table_start:
            print(f"Вторая таблица начинается: строка {second_table_start}")
        
        if revenue_row:
            print(f"Строка с прогнозом выручки: {revenue_row}")
            print(f"Формула в I{revenue_row}: {sheet[f'I{revenue_row}'].value}")
        
        if vat_row:
            print(f"Строка с НДС: {vat_row}")
            print(f"Формула в I{vat_row}: {sheet[f'I{vat_row}'].value}")
        
        workbook.close()
    except Exception as e:
        print(f"Ошибка при детальном просмотре: {str(e)}")

if __name__ == "__main__":
    process_budget_template()
    
    # Показываем детальный просмотр формул
    show_detailed_preview()