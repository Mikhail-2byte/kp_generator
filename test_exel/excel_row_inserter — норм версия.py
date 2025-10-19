from copy import copy
from pathlib import Path
import shutil

import openpyxl
from openpyxl.formula.translate import Translator


TEMPLATE_FILE = Path("Бюджет шаблон.xlsx")
OUTPUT_FILE = Path("Обработанный_добавление.xlsx")
DATA_START_ROW = 10
INSERT_ROW_INDEX = 11


def copy_row_style(sheet, source_row: int, target_row: int) -> None:
    """Копирует стили, формат чисел и высоту строки."""

    max_col = sheet.max_column
    for col in range(1, max_col + 1):
        src = sheet.cell(row=source_row, column=col)
        dst = sheet.cell(row=target_row, column=col)

        dst.value = None

        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)

    if source_row in sheet.row_dimensions:
        source_dim = sheet.row_dimensions[source_row]
        target_dim = sheet.row_dimensions[target_row]
        target_dim.height = source_dim.height


def copy_row_values(sheet, source_row: int, target_row: int) -> None:
    """Копирует значения и формулы, сдвигая относительные ссылки."""

    max_col = sheet.max_column

    for col in range(1, max_col + 1):
        src = sheet.cell(row=source_row, column=col)
        dst = sheet.cell(row=target_row, column=col)

        if src.value is None:
            dst.value = None
            continue

        if isinstance(src.value, str) and src.value.startswith("="):
            dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
        else:
            dst.value = src.value


def find_last_data_row(sheet) -> int:
    """Возвращает последнюю строку, в которой в столбце B указано число."""

    for row in range(sheet.max_row, DATA_START_ROW - 1, -1):
        cell_value = sheet[f"B{row}"].value
        if cell_value is None:
            continue
        try:
            int(cell_value)
            return row
        except (ValueError, TypeError):
            continue

    return DATA_START_ROW - 1


def update_total_row_formulas(sheet) -> None:
    """Пересчитывает формулы в итоговой строке и строке с НДС."""

    last_data_row = max(find_last_data_row(sheet), INSERT_ROW_INDEX)
    if last_data_row < DATA_START_ROW:
        return

    total_row = last_data_row + 1

    total_formulas = {
        "I": f"=SUM(I{DATA_START_ROW}:I{last_data_row})",
        "O": f"=SUM(O{DATA_START_ROW}:O{last_data_row})",
        "Q": f"=SUM(Q{DATA_START_ROW}:Q{last_data_row})",
        "S": f"=SUM(S{DATA_START_ROW}:S{last_data_row})",
        "U": f"=SUM(U{DATA_START_ROW}:U{last_data_row})",
        "Y": f"=SUM(Y{DATA_START_ROW}:Y{last_data_row})",
    }

    for col, formula in total_formulas.items():
        sheet[f"{col}{total_row}"] = formula

    sheet[f"X{total_row}"] = f"=AVERAGE(X{DATA_START_ROW}:X{last_data_row})"
    sheet[f"Z{total_row}"] = (
        f"=(I{total_row}-O{total_row}-S{total_row}-U{total_row}-"
        f"Y{total_row}-AA{total_row}-AC{total_row}-AB{total_row})/I{total_row}"
    )

    nds_row = total_row + 1
    sheet[f"I{nds_row}"] = f"=I{total_row}*1.2"


def update_summary_block(sheet) -> None:
    """Обновляет формулы итоговых строк c учётом новой строки."""

    last_data_row = max(find_last_data_row(sheet), INSERT_ROW_INDEX)
    total_row = last_data_row + 1
    sheet[f"I{total_row}"] = f"=SUM(I{DATA_START_ROW}:I{last_data_row})"
    sheet[f"I{total_row + 1}"] = f"=I{total_row}*1.2"

    k_formula_row = total_row + 4
    sheet[f"K{k_formula_row}"] = f"=I{k_formula_row}+I{k_formula_row + 1}"

    o_base_row = total_row
    sheet[f"O{o_base_row + 5}"] = f"=I{o_base_row}*14"
    sheet[f"O{o_base_row + 6}"] = f"=I{o_base_row}*14"

    i20_row = total_row + 8
    sheet[f"I{i20_row}"] = f"=I{total_row}"

    i25_row = total_row + 13
    sheet[f"I{i25_row}"] = f"=I{total_row + 1}"

    i26_row = i25_row + 1
    sheet[f"I{i26_row}"] = f"=I{i25_row}/120*20"

    i27_row = i26_row + 1
    sheet[f"I{i27_row}"] = f"=I{i25_row}-I{i26_row}"

    i28_row = i27_row + 1
    sheet[f"I{i28_row}"] = f"=SUM(I{i28_row + 1}:I{i28_row + 10})"

    i29_row = i28_row + 1
    sheet[f"I{i29_row}"] = f"=O{total_row}"

    i30_row = i29_row + 1
    d_reference_row = total_row + 32
    sheet[f"I{i30_row}"] = f"=IF(H{i30_row}=D{d_reference_row},I{i29_row}*3.2%,0)"

    i31_row = i30_row + 1
    sheet[f"I{i31_row}"] = f"=Y{total_row}"

    i32_row = i31_row + 1
    sheet[f"I{i32_row}"] = f"=S{total_row}"

    i33_row = i32_row + 1
    sheet[f"I{i33_row}"] = f"=U{total_row}"

    i34_row = i33_row + 1
    sheet[f"I{i34_row}"] = (
        f"=IF(H{i34_row}=\"ДА\",I{i29_row}*16%/365*K{k_formula_row},0)"
    )

    i35_row = i34_row + 1
    sheet[f"I{i35_row}"] = f"=AA{total_row}"

    i36_row = i35_row + 1
    sheet[f"I{i36_row}"] = f"=AB{total_row}"

    i37_row = i36_row + 1
    sheet[f"I{i37_row}"] = f"=AC{total_row}"

    i38_row = i37_row + 1
    sheet[f"I{i38_row}"] = (
        f"=IF(H{i38_row}=\"ДА\",I{i25_row}*3%/365*(I{total_row}+I{total_row + 1}),0)"
    )


def update_logistics_columns(sheet) -> None:
    """Настраивает формулы в столбцах R и T для всех строк с данными."""

    last_data_row = max(find_last_data_row(sheet), INSERT_ROW_INDEX)
    if last_data_row < DATA_START_ROW:
        return

    total_row = last_data_row + 1
    logistics_row = total_row + 3

    for row in range(DATA_START_ROW, last_data_row + 1):
        sheet[f"R{row}"] = f"=$U${logistics_row}/$Q${total_row}*P{row}/12*0.3"
        sheet[f"T{row}"] = f"=$U${logistics_row}/$Q${total_row}*P{row}/12*0.7"


def initialize_inserted_row(sheet) -> None:
    """Заполняет новую строку данными из строки-образца."""

    copy_row_values(sheet, DATA_START_ROW, INSERT_ROW_INDEX)


def insert_new_row() -> None:
    if not TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"Файл '{TEMPLATE_FILE}' не найден")

    shutil.copy2(TEMPLATE_FILE, OUTPUT_FILE)

    workbook = openpyxl.load_workbook(OUTPUT_FILE)
    sheet = workbook.active

    sheet.insert_rows(INSERT_ROW_INDEX)

    copy_row_style(sheet, DATA_START_ROW, INSERT_ROW_INDEX)
    initialize_inserted_row(sheet)

    update_total_row_formulas(sheet)
    update_summary_block(sheet)
    update_logistics_columns(sheet)

    workbook.save(OUTPUT_FILE)
    workbook.close()

    print(
        f"Добавлена новая строка {INSERT_ROW_INDEX}, форматы перенесены, итоговые "
        f"формулы обновлены. Результат сохранён в '{OUTPUT_FILE}'."
    )


if __name__ == "__main__":
    insert_new_row()
