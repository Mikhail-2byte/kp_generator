"""Тесты для маршрутов экспорта/импорта справочников логистики."""

import sys
from pathlib import Path

# Добавляем корень проекта в путь для возможности прямого запуска
project_root = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(project_root))

import pytest

from app.services import datasets


@pytest.fixture
def sample_main_cities_data(tmp_path: Path, app, monkeypatch):
    """Создает тестовые данные основных городов."""
    config_dir = tmp_path / 'config'
    config_dir.mkdir()
    
    original_config_dir = datasets.CONFIG_DIR
    monkeypatch.setattr(datasets, 'CONFIG_DIR', config_dir)
    
    cities = [
        {
            'name': 'Чита',
            'region': 'Забайкальский край',
            'truck_price': 50000.0,
            'is_main_route': True,
            'main_city': None
        }
    ]
    
    with app.app_context():
        datasets.save_main_cities(cities, actor='test_user')
        datasets.load_main_cities.cache_clear()
    
    yield config_dir
    
    monkeypatch.setattr(datasets, 'CONFIG_DIR', original_config_dir)


@pytest.fixture
def sample_ekb_rf_cities_data(tmp_path: Path, app, monkeypatch):
    """Создает тестовые данные городов ЕКБ+РФ."""
    config_dir = tmp_path / 'config'
    config_dir.mkdir()
    
    original_config_dir = datasets.CONFIG_DIR
    monkeypatch.setattr(datasets, 'CONFIG_DIR', config_dir)
    
    cities = [
        {
            'name': 'Москва',
            'region': 'Московская область',
            'distance_from_ekb_km': 1800
        }
    ]
    
    with app.app_context():
        datasets.save_ekb_rf_cities(cities, actor='test_user')
        datasets.load_ekb_rf_cities.cache_clear()
    
    yield config_dir
    
    monkeypatch.setattr(datasets, 'CONFIG_DIR', original_config_dir)


@pytest.fixture
def sample_trail_cities_data(tmp_path: Path, app, monkeypatch):
    """Создает тестовые данные городов трала."""
    config_dir = tmp_path / 'config'
    config_dir.mkdir()
    
    original_config_dir = datasets.CONFIG_DIR
    monkeypatch.setattr(datasets, 'CONFIG_DIR', config_dir)
    
    cities = [
        {
            'name': 'Новосибирск',
            'region': 'Новосибирская область',
            'trail_price': 120000.0
        }
    ]
    
    with app.app_context():
        datasets.save_trail_cities(cities, actor='test_user')
        datasets.load_trail_cities.cache_clear()
    
    yield config_dir
    
    monkeypatch.setattr(datasets, 'CONFIG_DIR', original_config_dir)


# Тесты экспорта основных городов
def test_export_main_cities(logged_in_client, sample_main_cities_data):
    """Тест экспорта основных городов в Excel."""
    response = logged_in_client.get('/admin/logistics/main/export')
    
    assert response.status_code == 200
    assert response.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    assert len(response.data) > 0
    content_disposition = response.headers.get('Content-Disposition', '')
    assert 'logistics_main_cities_' in content_disposition


def test_import_main_cities(logged_in_client, sample_main_cities_data, app, tmp_path):
    """Тест импорта основных городов из Excel."""
    # Создаем тестовый Excel файл
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl не установлен")
    
    wb = Workbook()
    ws = wb.active
    ws.append(['Город', 'Регион', 'Фура, руб.', 'Основной маршрут', 'Основной город'])
    ws.append(['Иркутск', 'Иркутская область', 60000, 'Да', ''])
    
    excel_file = tmp_path / 'test_import_main.xlsx'
    wb.save(excel_file)
    
    # Получаем CSRF токен
    page = logged_in_client.get('/admin/logistics?tab=main')
    assert page.status_code == 200
    csrf_token = page.data.decode().split('name="csrf_token" value="')[1].split('"')[0]
    
    # Импортируем
    with excel_file.open('rb') as f:
        response = logged_in_client.post('/admin/logistics/main/import', data={
            'excel_file': (f, 'test_import_main.xlsx'),
            'csrf_token': csrf_token
        }, content_type='multipart/form-data', follow_redirects=True)
    
    assert response.status_code == 200
    assert 'Импортировано городов' in response.data.decode('utf-8')
    
    # Проверяем, что данные импортированы
    with app.app_context():
        datasets.load_main_cities.cache_clear()
        cities = datasets.load_main_cities()
    
    assert len(cities) >= 1
    assert any(city['name'] == 'Иркутск' for city in cities)


def test_import_main_cities_invalid_file(logged_in_client, sample_main_cities_data):
    """Тест импорта основных городов с неверным форматом файла."""
    # Получаем CSRF токен
    page = logged_in_client.get('/admin/logistics?tab=main')
    assert page.status_code == 200
    csrf_token = page.data.decode().split('name="csrf_token" value="')[1].split('"')[0]
    
    # Пытаемся импортировать текстовый файл
    response = logged_in_client.post('/admin/logistics/main/import', data={
        'excel_file': (b'not an excel file', 'test.txt'),
        'csrf_token': csrf_token
    }, content_type='multipart/form-data', follow_redirects=True)
    
    assert response.status_code == 200
    # Проверяем, что произошел редирект на страницу логистики (не остались на странице импорта)
    assert '/admin/logistics' in response.request.path or 'admin_logistics' in response.data.decode('utf-8')


def test_import_main_cities_no_file(logged_in_client, sample_main_cities_data):
    """Тест импорта основных городов без файла."""
    # Получаем CSRF токен
    page = logged_in_client.get('/admin/logistics?tab=main')
    assert page.status_code == 200
    csrf_token = page.data.decode().split('name="csrf_token" value="')[1].split('"')[0]
    
    response = logged_in_client.post('/admin/logistics/main/import', data={
        'csrf_token': csrf_token
    }, follow_redirects=True)
    
    assert response.status_code == 200
    assert 'Файл не выбран' in response.data.decode('utf-8')


# Тесты экспорта городов ЕКБ+РФ
def test_export_ekb_rf_cities(logged_in_client, sample_ekb_rf_cities_data):
    """Тест экспорта городов ЕКБ+РФ в Excel."""
    response = logged_in_client.get('/admin/logistics/ekb-rf/export')
    
    assert response.status_code == 200
    assert response.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    assert len(response.data) > 0
    content_disposition = response.headers.get('Content-Disposition', '')
    assert 'logistics_ekb_rf_cities_' in content_disposition


def test_import_ekb_rf_cities(logged_in_client, sample_ekb_rf_cities_data, app, tmp_path):
    """Тест импорта городов ЕКБ+РФ из Excel."""
    # Создаем тестовый Excel файл
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl не установлен")
    
    wb = Workbook()
    ws = wb.active
    ws.append(['Город', 'Регион', 'Расстояние от ЕКБ, км'])
    ws.append(['Санкт-Петербург', 'Ленинградская область', 2100])
    
    excel_file = tmp_path / 'test_import_ekb_rf.xlsx'
    wb.save(excel_file)
    
    # Получаем CSRF токен
    page = logged_in_client.get('/admin/logistics?tab=ekb_rf')
    assert page.status_code == 200
    csrf_token = page.data.decode().split('name="csrf_token" value="')[1].split('"')[0]
    
    # Импортируем
    with excel_file.open('rb') as f:
        response = logged_in_client.post('/admin/logistics/ekb-rf/import', data={
            'excel_file': (f, 'test_import_ekb_rf.xlsx'),
            'csrf_token': csrf_token
        }, content_type='multipart/form-data', follow_redirects=True)
    
    assert response.status_code == 200
    assert 'Импортировано городов' in response.data.decode('utf-8')
    
    # Проверяем, что данные импортированы
    with app.app_context():
        datasets.load_ekb_rf_cities.cache_clear()
        cities = datasets.load_ekb_rf_cities()
    
    assert len(cities) >= 1
    assert any(city['name'] == 'Санкт-Петербург' for city in cities)


# Тесты экспорта городов трала
def test_export_trail_cities(logged_in_client, sample_trail_cities_data):
    """Тест экспорта городов трала в Excel."""
    response = logged_in_client.get('/admin/logistics/trail/export')
    
    assert response.status_code == 200
    assert response.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    assert len(response.data) > 0
    content_disposition = response.headers.get('Content-Disposition', '')
    assert 'logistics_trail_cities_' in content_disposition


def test_import_trail_cities(logged_in_client, sample_trail_cities_data, app, tmp_path):
    """Тест импорта городов трала из Excel."""
    # Создаем тестовый Excel файл
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl не установлен")
    
    wb = Workbook()
    ws = wb.active
    ws.append(['Город', 'Регион', 'Цена трала, руб.'])
    ws.append(['Красноярск', 'Красноярский край', 100000])
    
    excel_file = tmp_path / 'test_import_trail.xlsx'
    wb.save(excel_file)
    
    # Получаем CSRF токен
    page = logged_in_client.get('/admin/logistics?tab=trail')
    assert page.status_code == 200
    csrf_token = page.data.decode().split('name="csrf_token" value="')[1].split('"')[0]
    
    # Импортируем
    with excel_file.open('rb') as f:
        response = logged_in_client.post('/admin/logistics/trail/import', data={
            'excel_file': (f, 'test_import_trail.xlsx'),
            'csrf_token': csrf_token
        }, content_type='multipart/form-data', follow_redirects=True)
    
    assert response.status_code == 200
    assert 'Импортировано городов' in response.data.decode('utf-8')
    
    # Проверяем, что данные импортированы
    with app.app_context():
        datasets.load_trail_cities.cache_clear()
        cities = datasets.load_trail_cities()
    
    assert len(cities) >= 1
    assert any(city['name'] == 'Красноярск' for city in cities)


# Тесты на доступность маршрутов
def test_export_main_cities_requires_admin(client):
    """Тест, что экспорт основных городов требует прав администратора."""
    response = client.get('/admin/logistics/main/export', follow_redirects=True)
    
    # Должно быть перенаправление на страницу входа
    assert response.status_code == 200
    assert 'login' in response.request.path.lower() or 'Войти' in response.data.decode('utf-8')


def test_export_ekb_rf_cities_requires_admin(client):
    """Тест, что экспорт городов ЕКБ+РФ требует прав администратора."""
    response = client.get('/admin/logistics/ekb-rf/export', follow_redirects=True)
    
    # Должно быть перенаправление на страницу входа
    assert response.status_code == 200
    assert 'login' in response.request.path.lower() or 'Войти' in response.data.decode('utf-8')


def test_export_trail_cities_requires_admin(client):
    """Тест, что экспорт городов трала требует прав администратора."""
    response = client.get('/admin/logistics/trail/export', follow_redirects=True)
    
    # Должно быть перенаправление на страницу входа
    assert response.status_code == 200
    assert 'login' in response.request.path.lower() or 'Войти' in response.data.decode('utf-8')


if __name__ == "__main__":
    pytest.main([__file__, "-v"])

