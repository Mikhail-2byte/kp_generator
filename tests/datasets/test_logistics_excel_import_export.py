"""Тесты для экспорта и импорта справочников логистики в Excel."""

import sys
from pathlib import Path

# Добавляем корень проекта в путь для возможности прямого запуска
project_root = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(project_root))

import pytest

from app.services import datasets


@pytest.fixture
def sample_main_cities():
    """Создает тестовые данные основных городов."""
    return [
        {
            'name': 'Чита',
            'region': 'Забайкальский край',
            'truck_price': 50000.0,
            'is_main_route': True,
            'main_city': None
        },
        {
            'name': 'Краснокаменск',
            'region': 'Забайкальский край',
            'truck_price': 45000.0,
            'is_main_route': False,
            'main_city': 'Чита'
        }
    ]


@pytest.fixture
def sample_ekb_rf_cities():
    """Создает тестовые данные городов ЕКБ+РФ."""
    return [
        {
            'name': 'Москва',
            'region': 'Московская область',
            'distance_from_ekb_km': 1800
        },
        {
            'name': 'Санкт-Петербург',
            'region': 'Ленинградская область',
            'distance_from_ekb_km': 2100
        }
    ]


@pytest.fixture
def sample_trail_cities():
    """Создает тестовые данные городов трала."""
    return [
        {
            'name': 'Новосибирск',
            'region': 'Новосибирская область',
            'trail_price': 120000.0
        },
        {
            'name': 'Красноярск',
            'region': 'Красноярский край',
            'trail_price': 100000.0
        }
    ]


# Тесты для основных городов
def test_export_main_cities_to_excel(tmp_path: Path, sample_main_cities, app, monkeypatch):
    """Тест экспорта основных городов в Excel."""
    config_dir = tmp_path / 'config'
    config_dir.mkdir()
    
    original_config_dir = datasets.CONFIG_DIR
    monkeypatch.setattr(datasets, 'CONFIG_DIR', config_dir)
    
    try:
        # Сохраняем тестовые данные
        with app.app_context():
            datasets.save_main_cities(sample_main_cities, actor='test_user')
            datasets.load_main_cities.cache_clear()
            
            # Экспортируем
            excel_data = datasets.export_main_cities_to_excel()
        
        # Проверяем, что данные не пустые
        assert len(excel_data) > 0
        
        # Сохраняем во временный файл для проверки
        excel_file = tmp_path / 'exported_main.xlsx'
        excel_file.write_bytes(excel_data)
        
        # Проверяем, что файл создан
        assert excel_file.exists()
        assert excel_file.stat().st_size > 0
        
    finally:
        monkeypatch.setattr(datasets, 'CONFIG_DIR', original_config_dir)


def test_import_main_cities_from_excel(tmp_path: Path, app, monkeypatch):
    """Тест импорта основных городов из Excel."""
    config_dir = tmp_path / 'config'
    config_dir.mkdir()
    
    original_config_dir = datasets.CONFIG_DIR
    monkeypatch.setattr(datasets, 'CONFIG_DIR', config_dir)
    
    try:
        # Создаем Excel файл вручную с помощью openpyxl
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl не установлен")
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Город', 'Регион', 'Фура, руб.', 'Основной маршрут', 'Основной город'])
        ws.append(['Чита', 'Забайкальский край', 50000, 'Да', ''])
        ws.append(['Краснокаменск', 'Забайкальский край', 45000, 'Нет', 'Чита'])
        
        excel_file = tmp_path / 'test_main.xlsx'
        wb.save(excel_file)
        
        # Импортируем
        with app.app_context():
            count = datasets.import_main_cities_from_excel(excel_file, actor='test_user')
        
        assert count == 2
        
        # Проверяем данные
        datasets.load_main_cities.cache_clear()
        with app.app_context():
            cities = datasets.load_main_cities()
        
        assert len(cities) == 2
        assert cities[0]['name'] == 'Чита'
        assert cities[0]['is_main_route'] is True
        assert cities[1]['name'] == 'Краснокаменск'
        assert cities[1]['is_main_route'] is False
        assert cities[1]['main_city'] == 'Чита'
        
    finally:
        monkeypatch.setattr(datasets, 'CONFIG_DIR', original_config_dir)


def test_import_main_cities_from_excel_with_bool_values(tmp_path: Path, app, monkeypatch):
    """Тест импорта основных городов с различными форматами булевых значений."""
    config_dir = tmp_path / 'config'
    config_dir.mkdir()
    
    original_config_dir = datasets.CONFIG_DIR
    monkeypatch.setattr(datasets, 'CONFIG_DIR', config_dir)
    
    try:
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl не установлен")
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Город', 'Регион', 'Фура, руб.', 'Основной маршрут', 'Основной город'])
        ws.append(['Город1', 'Регион1', 50000, True, ''])  # Булево значение
        ws.append(['Город2', 'Регион2', 45000, 'да', 'Город1'])  # Строка "да"
        ws.append(['Город3', 'Регион3', 40000, 1, 'Город1'])  # Число 1
        
        excel_file = tmp_path / 'test_main_bool.xlsx'
        wb.save(excel_file)
        
        # Импортируем
        with app.app_context():
            count = datasets.import_main_cities_from_excel(excel_file, actor='test_user')
        
        assert count == 3
        
        # Проверяем данные
        datasets.load_main_cities.cache_clear()
        with app.app_context():
            cities = datasets.load_main_cities()
        
        assert cities[0]['is_main_route'] is True
        assert cities[1]['is_main_route'] is True
        assert cities[2]['is_main_route'] is True
        
    finally:
        monkeypatch.setattr(datasets, 'CONFIG_DIR', original_config_dir)


def test_export_main_cities_empty(tmp_path: Path, app, monkeypatch):
    """Тест экспорта пустого справочника основных городов."""
    config_dir = tmp_path / 'config'
    config_dir.mkdir()
    
    original_config_dir = datasets.CONFIG_DIR
    monkeypatch.setattr(datasets, 'CONFIG_DIR', config_dir)
    
    try:
        with app.app_context():
            datasets.save_main_cities([], actor='test_user')
            datasets.load_main_cities.cache_clear()
            
            excel_data = datasets.export_main_cities_to_excel()
        
        # Даже пустой справочник должен создавать файл с заголовками
        assert len(excel_data) > 0
        
    finally:
        monkeypatch.setattr(datasets, 'CONFIG_DIR', original_config_dir)


# Тесты для городов ЕКБ+РФ
def test_export_ekb_rf_cities_to_excel(tmp_path: Path, sample_ekb_rf_cities, app, monkeypatch):
    """Тест экспорта городов ЕКБ+РФ в Excel."""
    config_dir = tmp_path / 'config'
    config_dir.mkdir()
    
    original_config_dir = datasets.CONFIG_DIR
    monkeypatch.setattr(datasets, 'CONFIG_DIR', config_dir)
    
    try:
        # Сохраняем тестовые данные
        with app.app_context():
            datasets.save_ekb_rf_cities(sample_ekb_rf_cities, actor='test_user')
            datasets.load_ekb_rf_cities.cache_clear()
            
            # Экспортируем
            excel_data = datasets.export_ekb_rf_cities_to_excel()
        
        # Проверяем, что данные не пустые
        assert len(excel_data) > 0
        
        # Сохраняем во временный файл для проверки
        excel_file = tmp_path / 'exported_ekb_rf.xlsx'
        excel_file.write_bytes(excel_data)
        
        # Проверяем, что файл создан
        assert excel_file.exists()
        assert excel_file.stat().st_size > 0
        
    finally:
        monkeypatch.setattr(datasets, 'CONFIG_DIR', original_config_dir)


def test_import_ekb_rf_cities_from_excel(tmp_path: Path, app, monkeypatch):
    """Тест импорта городов ЕКБ+РФ из Excel."""
    config_dir = tmp_path / 'config'
    config_dir.mkdir()
    
    original_config_dir = datasets.CONFIG_DIR
    monkeypatch.setattr(datasets, 'CONFIG_DIR', config_dir)
    
    try:
        # Создаем Excel файл вручную
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl не установлен")
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Город', 'Регион', 'Расстояние от ЕКБ, км'])
        ws.append(['Москва', 'Московская область', 1800])
        ws.append(['Санкт-Петербург', 'Ленинградская область', 2100])
        ws.append(['Город без расстояния', 'Регион', None])  # Пустое расстояние
        
        excel_file = tmp_path / 'test_ekb_rf.xlsx'
        wb.save(excel_file)
        
        # Импортируем
        with app.app_context():
            count = datasets.import_ekb_rf_cities_from_excel(excel_file, actor='test_user')
        
        assert count == 3
        
        # Проверяем данные
        datasets.load_ekb_rf_cities.cache_clear()
        with app.app_context():
            cities = datasets.load_ekb_rf_cities()
        
        assert len(cities) == 3
        assert cities[0]['name'] == 'Москва'
        assert cities[0]['distance_from_ekb_km'] == 1800
        assert cities[2]['name'] == 'Город без расстояния'
        assert cities[2]['distance_from_ekb_km'] is None
        
    finally:
        monkeypatch.setattr(datasets, 'CONFIG_DIR', original_config_dir)


def test_import_ekb_rf_cities_from_excel_empty_distance(tmp_path: Path, app, monkeypatch):
    """Тест импорта городов ЕКБ+РФ с пустым расстоянием."""
    config_dir = tmp_path / 'config'
    config_dir.mkdir()
    
    original_config_dir = datasets.CONFIG_DIR
    monkeypatch.setattr(datasets, 'CONFIG_DIR', config_dir)
    
    try:
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl не установлен")
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Город', 'Регион', 'Расстояние от ЕКБ, км'])
        ws.append(['Город1', 'Регион1', ''])  # Пустая строка
        ws.append(['Город2', 'Регион2', None])  # None
        
        excel_file = tmp_path / 'test_ekb_rf_empty.xlsx'
        wb.save(excel_file)
        
        # Импортируем
        with app.app_context():
            count = datasets.import_ekb_rf_cities_from_excel(excel_file, actor='test_user')
        
        assert count == 2
        
        # Проверяем данные
        datasets.load_ekb_rf_cities.cache_clear()
        with app.app_context():
            cities = datasets.load_ekb_rf_cities()
        
        assert cities[0]['distance_from_ekb_km'] is None
        assert cities[1]['distance_from_ekb_km'] is None
        
    finally:
        monkeypatch.setattr(datasets, 'CONFIG_DIR', original_config_dir)


# Тесты для городов трала
def test_export_trail_cities_to_excel(tmp_path: Path, sample_trail_cities, app, monkeypatch):
    """Тест экспорта городов трала в Excel."""
    config_dir = tmp_path / 'config'
    config_dir.mkdir()
    
    original_config_dir = datasets.CONFIG_DIR
    monkeypatch.setattr(datasets, 'CONFIG_DIR', config_dir)
    
    try:
        # Сохраняем тестовые данные
        with app.app_context():
            datasets.save_trail_cities(sample_trail_cities, actor='test_user')
            datasets.load_trail_cities.cache_clear()
            
            # Экспортируем
            excel_data = datasets.export_trail_cities_to_excel()
        
        # Проверяем, что данные не пустые
        assert len(excel_data) > 0
        
        # Сохраняем во временный файл для проверки
        excel_file = tmp_path / 'exported_trail.xlsx'
        excel_file.write_bytes(excel_data)
        
        # Проверяем, что файл создан
        assert excel_file.exists()
        assert excel_file.stat().st_size > 0
        
    finally:
        monkeypatch.setattr(datasets, 'CONFIG_DIR', original_config_dir)


def test_import_trail_cities_from_excel(tmp_path: Path, app, monkeypatch):
    """Тест импорта городов трала из Excel."""
    config_dir = tmp_path / 'config'
    config_dir.mkdir()
    
    original_config_dir = datasets.CONFIG_DIR
    monkeypatch.setattr(datasets, 'CONFIG_DIR', config_dir)
    
    try:
        # Создаем Excel файл вручную
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl не установлен")
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Город', 'Регион', 'Цена трала, руб.'])
        ws.append(['Новосибирск', 'Новосибирская область', 120000])
        ws.append(['Красноярск', 'Красноярский край', 100000])
        
        excel_file = tmp_path / 'test_trail.xlsx'
        wb.save(excel_file)
        
        # Импортируем
        with app.app_context():
            count = datasets.import_trail_cities_from_excel(excel_file, actor='test_user')
        
        assert count == 2
        
        # Проверяем данные
        datasets.load_trail_cities.cache_clear()
        with app.app_context():
            cities = datasets.load_trail_cities()
        
        assert len(cities) == 2
        assert cities[0]['name'] == 'Новосибирск'
        assert cities[0]['trail_price'] == 120000.0
        assert cities[1]['name'] == 'Красноярск'
        assert cities[1]['trail_price'] == 100000.0
        
    finally:
        monkeypatch.setattr(datasets, 'CONFIG_DIR', original_config_dir)


def test_import_trail_cities_from_excel_with_string_price(tmp_path: Path, app, monkeypatch):
    """Тест импорта городов трала со строковым значением цены."""
    config_dir = tmp_path / 'config'
    config_dir.mkdir()
    
    original_config_dir = datasets.CONFIG_DIR
    monkeypatch.setattr(datasets, 'CONFIG_DIR', config_dir)
    
    try:
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl не установлен")
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Город', 'Регион', 'Цена трала, руб.'])
        ws.append(['Город1', 'Регион1', '120000'])  # Строка
        ws.append(['Город2', 'Регион2', '100500'])  # Строка без запятой (запятая интерпретируется как десятичный разделитель)
        
        excel_file = tmp_path / 'test_trail_string.xlsx'
        wb.save(excel_file)
        
        # Импортируем
        with app.app_context():
            count = datasets.import_trail_cities_from_excel(excel_file, actor='test_user')
        
        assert count == 2
        
        # Проверяем данные
        datasets.load_trail_cities.cache_clear()
        with app.app_context():
            cities = datasets.load_trail_cities()
        
        assert cities[0]['trail_price'] == 120000.0
        assert cities[1]['trail_price'] == 100500.0
        
    finally:
        monkeypatch.setattr(datasets, 'CONFIG_DIR', original_config_dir)


# Тесты на обработку ошибок
def test_import_main_cities_from_excel_nonexistent_file(tmp_path: Path, app):
    """Тест импорта основных городов из несуществующего файла."""
    nonexistent_file = tmp_path / 'nonexistent.xlsx'
    
    with app.app_context():
        with pytest.raises(FileNotFoundError):
            datasets.import_main_cities_from_excel(nonexistent_file, actor='test_user')


def test_import_ekb_rf_cities_from_excel_nonexistent_file(tmp_path: Path, app):
    """Тест импорта городов ЕКБ+РФ из несуществующего файла."""
    nonexistent_file = tmp_path / 'nonexistent.xlsx'
    
    with app.app_context():
        with pytest.raises(FileNotFoundError):
            datasets.import_ekb_rf_cities_from_excel(nonexistent_file, actor='test_user')


def test_import_trail_cities_from_excel_nonexistent_file(tmp_path: Path, app):
    """Тест импорта городов трала из несуществующего файла."""
    nonexistent_file = tmp_path / 'nonexistent.xlsx'
    
    with app.app_context():
        with pytest.raises(FileNotFoundError):
            datasets.import_trail_cities_from_excel(nonexistent_file, actor='test_user')


def test_import_main_cities_from_excel_invalid_data(tmp_path: Path, app, monkeypatch):
    """Тест импорта основных городов с невалидными данными."""
    config_dir = tmp_path / 'config'
    config_dir.mkdir()
    
    original_config_dir = datasets.CONFIG_DIR
    monkeypatch.setattr(datasets, 'CONFIG_DIR', config_dir)
    
    try:
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl не установлен")
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Город', 'Регион', 'Фура, руб.', 'Основной маршрут', 'Основной город'])
        ws.append(['', 'Регион', 50000, 'Да', ''])  # Пустое имя города - должно быть пропущено
        ws.append(['Город1', 'Регион1', 'не число', 'Да', ''])  # Невалидная цена
        
        excel_file = tmp_path / 'test_main_invalid.xlsx'
        wb.save(excel_file)
        
        # Импортируем
        with app.app_context():
            count = datasets.import_main_cities_from_excel(excel_file, actor='test_user')
        
        # Должна быть импортирована только одна запись (с валидным именем)
        assert count == 1
        
    finally:
        monkeypatch.setattr(datasets, 'CONFIG_DIR', original_config_dir)


def test_import_ekb_rf_cities_from_excel_minimal_columns(tmp_path: Path, app, monkeypatch):
    """Тест импорта городов ЕКБ+РФ с минимальным количеством колонок."""
    config_dir = tmp_path / 'config'
    config_dir.mkdir()
    
    original_config_dir = datasets.CONFIG_DIR
    monkeypatch.setattr(datasets, 'CONFIG_DIR', config_dir)
    
    try:
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl не установлен")
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Город', 'Регион'])  # Без колонки расстояния
        ws.append(['Город1', 'Регион1'])
        
        excel_file = tmp_path / 'test_ekb_rf_min.xlsx'
        wb.save(excel_file)
        
        # Импортируем
        with app.app_context():
            count = datasets.import_ekb_rf_cities_from_excel(excel_file, actor='test_user')
        
        assert count == 1
        
        # Проверяем данные
        datasets.load_ekb_rf_cities.cache_clear()
        with app.app_context():
            cities = datasets.load_ekb_rf_cities()
        
        assert cities[0]['distance_from_ekb_km'] is None
        
    finally:
        monkeypatch.setattr(datasets, 'CONFIG_DIR', original_config_dir)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])

