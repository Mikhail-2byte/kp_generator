from io import BytesIO

import pytest
from openpyxl import Workbook

from app.services.excel_importer import ExcelImportError, ExcelImporterService


def _build_workbook(header, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _build_fixed_template_workbook(positions_data):
    """
    Создает Excel файл с новым форматом (фиксированные ячейки начиная с 7-й строки).
    
    positions_data: список словарей с ключами:
        - product (B)
        - material (D)
        - drawing_number (E)
        - quantity (K)
        - cost_price (L)
        - total_cost (M) - опционально
        - weight (N)
        - duty_percent (P) - опционально
        - cost_price_per_kg (S) - опционально
    """
    wb = Workbook()
    ws = wb.active
    
    # Заполняем пустые строки до 7-й (для реалистичности)
    for i in range(1, 7):
        ws.append([''] * 20)
    
    # Заполняем данные начиная с 7-й строки
    for pos_data in positions_data:
        row_num = len([r for r in ws.iter_rows(min_row=1, max_row=6)]) + len([r for r in ws.iter_rows(min_row=7)]) + 1
        if row_num < 7:
            row_num = 7
        
        # B (колонка 2) - product
        ws.cell(row=row_num, column=2, value=pos_data.get('product', ''))
        # D (колонка 4) - material
        ws.cell(row=row_num, column=4, value=pos_data.get('material', ''))
        # E (колонка 5) - drawing_number
        ws.cell(row=row_num, column=5, value=pos_data.get('drawing_number', ''))
        # K (колонка 11) - quantity
        ws.cell(row=row_num, column=11, value=pos_data.get('quantity', ''))
        # L (колонка 12) - cost_price
        ws.cell(row=row_num, column=12, value=pos_data.get('cost_price', ''))
        # M (колонка 13) - total_cost
        ws.cell(row=row_num, column=13, value=pos_data.get('total_cost', ''))
        # N (колонка 14) - weight
        ws.cell(row=row_num, column=14, value=pos_data.get('weight', ''))
        # P (колонка 16) - duty_percent
        ws.cell(row=row_num, column=16, value=pos_data.get('duty_percent', ''))
        # S (колонка 19) - cost_price_per_kg
        ws.cell(row=row_num, column=19, value=pos_data.get('cost_price_per_kg', ''))
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def test_excel_importer_parses_positions():
    """Тест на старый формат (по заголовкам)."""
    header = ['Номенклатура', 'Цена закупа', 'Количество', 'Вес за шт. (кг)', 'Пошлина (%)']
    rows = [
        ['Труба', '1200', '5', '1.5', '7'],
        ['Фланец', '800', '2', '0.8', ''],
    ]
    stream = _build_workbook(header, rows)
    importer = ExcelImporterService()

    positions = importer.parse_positions(stream)

    assert len(positions) == 2
    assert positions[0]['product'] == 'Труба'
    assert positions[1]['duty_percent'] == '0'


def test_excel_importer_missing_columns_raises():
    """Тест на ошибку при отсутствии обязательных колонок в старом формате."""
    header = ['Номенклатура', 'Количество']
    rows = [['Труба', '5']]
    stream = _build_workbook(header, rows)
    importer = ExcelImporterService()

    with pytest.raises(ExcelImportError):
        importer.parse_positions(stream)


def test_excel_importer_parses_fixed_template():
    """Тест на новый формат (фиксированные ячейки начиная с 7-й строки)."""
    positions_data = [
        {
            'product': 'Труба стальная',
            'material': 'Сталь',
            'drawing_number': 'Ч-001',
            'quantity': '10',
            'cost_price': '1200',
            'total_cost': '12000',
            'weight': '1.5',
            'duty_percent': '7',
        },
        {
            'product': 'Фланец',
            'material': 'Чугун',
            'drawing_number': 'Ч-002',
            'quantity': '5',
            'cost_price': '800',
            'total_cost': '4000',
            'weight': '0.8',
            'duty_percent': '',
        },
    ]
    stream = _build_fixed_template_workbook(positions_data)
    importer = ExcelImporterService()

    positions = importer.parse_positions(stream)

    assert len(positions) == 2
    assert positions[0]['product'] == 'Труба стальная'
    assert positions[0]['material'] == 'Сталь'
    assert positions[0]['drawing_number'] == 'Ч-001'
    assert positions[0]['quantity'] == '10'
    assert positions[0]['weight'] == '1.5'
    assert positions[0]['duty_percent'] == '7'
    assert 'cost_price' in positions[0]
    assert 'cost_price_per_kg' in positions[0]
    
    assert positions[1]['product'] == 'Фланец'
    assert positions[1]['duty_percent'] == '0'  # Пустое значение должно стать '0'


def test_excel_importer_fixed_template_stops_on_empty_row():
    """Тест, что пустые строки после данных не создают лишние позиции."""
    positions_data = [
        {
            'product': 'Труба',
            'quantity': '5',
            'cost_price': '1200',
            'weight': '1.5',
        },
        {
            'product': 'Фланец',
            'quantity': '2',
            'cost_price': '800',
            'weight': '0.8',
        },
        # Пустая строка (все поля пустые)
        {
            'product': '',
            'quantity': '',
            'weight': '',
        },
        # Еще одна пустая строка
        {
            'product': '',
            'quantity': '',
            'weight': '',
        },
    ]
    stream = _build_fixed_template_workbook(positions_data)
    importer = ExcelImporterService()

    positions = importer.parse_positions(stream)

    # Должны быть только 2 позиции, пустые строки должны игнорироваться
    assert len(positions) == 2
    assert positions[0]['product'] == 'Труба'
    assert positions[1]['product'] == 'Фланец'


def test_excel_importer_fixed_template_ignores_footer_with_totals():
    """Новый формат: строки с итогами (без product и quantity) игнорируются, даже если есть вес/сумма."""
    positions_data = [
        {
            'product': 'Труба',
            'quantity': '5',
            'cost_price': '1200',
            'weight': '1.5',
        },
        {
            'product': 'Фланец',
            'quantity': '2',
            'cost_price': '800',
            'weight': '0.8',
        },
        # Строка итогов: нет product и quantity, но есть total_cost и общий вес
        {
            'product': '',
            'quantity': '',
            'weight': '10.0',      # общий вес
            'total_cost': '20000',  # общая сумма
        },
    ]
    stream = _build_fixed_template_workbook(positions_data)
    importer = ExcelImporterService()

    positions = importer.parse_positions(stream)

    # Итоговая строка не должна превращаться в позицию
    assert len(positions) == 2
    assert positions[0]['product'] == 'Труба'
    assert positions[1]['product'] == 'Фланец'


def test_excel_importer_fallback_to_fixed_template():
    """Тест на переключение: если заголовков нет, используется фиксированный шаблон."""
    # Создаем файл без заголовков, но с данными в фиксированных ячейках
    positions_data = [
        {
            'product': 'Товар без заголовков',
            'quantity': '3',
            'cost_price': '1000',
            'weight': '2.0',
        },
    ]
    stream = _build_fixed_template_workbook(positions_data)
    importer = ExcelImporterService()

    positions = importer.parse_positions(stream)

    assert len(positions) == 1
    assert positions[0]['product'] == 'Товар без заголовков'
    assert positions[0]['quantity'] == '3'


def test_excel_importer_fixed_template_requires_mandatory_fields():
    """Тест, что новый формат требует обязательные поля."""
    positions_data = [
        {
            # Отсутствует product
            'product': '',
            'quantity': '5',
            'cost_price': '1200',
            'weight': '1.5',
        },
    ]
    stream = _build_fixed_template_workbook(positions_data)
    importer = ExcelImporterService()

    with pytest.raises(ExcelImportError) as exc_info:
        importer.parse_positions(stream)
    
    assert 'отсутствует значение' in str(exc_info.value).lower() or 'обязательные' in str(exc_info.value).lower()


def test_excel_importer_fixed_template_with_cost_price_per_kg():
    """Новый формат: указаны и цена за шт. (L), и цена за кг (S) — значения согласованы."""
    positions_data = [
        {
            'product': 'Труба',
            'quantity': '4',
            'weight': '2.5',           # 2.5 кг за шт.
            'cost_price_per_kg': '100',  # 100 руб/кг
            # Ожидаемая цена за шт. = 100 * 2.5 = 250
            'cost_price': '250',
        },
    ]
    stream = _build_fixed_template_workbook(positions_data)
    importer = ExcelImporterService()

    positions = importer.parse_positions(stream)

    assert len(positions) == 1
    pos = positions[0]
    # Должны получить те же значения, что указаны в файле
    assert pos['cost_price'] == '250'
    assert pos['cost_price_per_kg'] == '100'


def test_excel_importer_fixed_template_only_cost_price_per_kg():
    """Новый формат: указана только цена за кг (S), цена за шт. вычисляется по весу."""
    positions_data = [
        {
            'product': 'Труба',
            'quantity': '3',
            'weight': '2.0',             # 2 кг за шт.
            'cost_price_per_kg': '150',  # 150 руб/кг
            # cost_price не указываем — должна вычислиться: 150 * 2 = 300
        },
    ]
    stream = _build_fixed_template_workbook(positions_data)
    importer = ExcelImporterService()

    positions = importer.parse_positions(stream)

    assert len(positions) == 1
    pos = positions[0]
    assert pos['cost_price_per_kg'] == '150'
    # 150 * 2.0 = 300
    assert pos['cost_price'] == '300'


def test_excel_importer_fixed_template_inconsistent_cost_price_and_per_kg():
    """Новый формат: несогласованные L и S должны приводить к ошибке."""
    positions_data = [
        {
            'product': 'Труба',
            'quantity': '2',
            'weight': '1.0',
            'cost_price_per_kg': '100',  # 100 руб/кг
            # Ожидаемая цена за шт. = 100 * 1 = 100, но укажем другое значение
            'cost_price': '200',
        },
    ]
    stream = _build_fixed_template_workbook(positions_data)
    importer = ExcelImporterService()

    with pytest.raises(ExcelImportError):
        importer.parse_positions(stream)

